"""
Excel Export Service

This service handles all Excel export functionality with reusable,
configurable export methods.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict, Any, Optional
from django.db.models import QuerySet


class ExcelExportService:
    """Service class for generating Excel exports"""

    # Style constants
    HEADER_FILL = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
    TITLE_FONT = Font(bold=True, size=14)
    BORDER_THIN = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def __init__(self, title: str = "Export"):
        """Initialize export service"""
        self.workbook = openpyxl.Workbook()
        self.title = title

    def create_worksheet(self, title: str, headers: List[str], data: List[Dict[str, Any]],
                        column_widths: Optional[Dict[str, int]] = None) -> None:
        """
        Create a worksheet with data

        Args:
            title: Worksheet title
            headers: List of column headers
            data: List of dictionaries with row data
            column_widths: Optional dictionary mapping column letters to widths
        """
        # Remove default sheet if this is first custom sheet
        if 'Sheet' in self.workbook.sheetnames and len(self.workbook.sheetnames) == 1:
            del self.workbook['Sheet']

        ws = self.workbook.create_sheet(title=title)

        # Add title
        ws.append([self.title])
        ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
        ws['A1'].font = self.TITLE_FONT
        ws['A1'].alignment = Alignment(horizontal='center')

        # Add generation timestamp
        ws.append([f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
        ws.append([])  # Empty row

        # Add headers
        header_row = len(ws['A']) + 1
        ws.append(headers)

        # Style headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER_THIN

        # Add data rows
        for row_data in data:
            row = [row_data.get(header, '') for header in headers]
            ws.append(row)

            # Apply borders to data cells
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=ws.max_row, column=col_num)
                cell.border = self.BORDER_THIN

        # Auto-adjust column widths or apply custom widths
        if column_widths:
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
        else:
            self._auto_adjust_columns(ws, headers)

        # Freeze header row
        ws.freeze_panes = f'A{header_row + 1}'

    def _auto_adjust_columns(self, ws, headers: List[str]) -> None:
        """Auto-adjust column widths based on content"""
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            max_length = len(str(header))

            # Check data cells
            for row in range(4, ws.max_row + 1):  # Start after title and headers
                cell_value = ws.cell(row=row, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            # Set width with padding
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    def add_summary_sheet(self, summary_data: Dict[str, Any]) -> None:
        """
        Add a summary sheet with key metrics

        Args:
            summary_data: Dictionary of key-value pairs for summary
        """
        ws = self.workbook.create_sheet(title='Summary', index=0)

        # Title
        ws.append([f'{self.title} - Summary'])
        ws['A1'].font = self.TITLE_FONT
        ws.merge_cells('A1:B1')

        ws.append([])

        # Add summary data
        for key, value in summary_data.items():
            ws.append([key, value])
            ws[f'A{ws.max_row}'].font = Font(bold=True)

        # Style
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

    def save(self, filepath: str) -> str:
        """
        Save workbook to file

        Args:
            filepath: Path to save the Excel file

        Returns:
            str: The filepath where file was saved
        """
        self.workbook.save(filepath)
        return filepath

    def get_workbook(self) -> openpyxl.Workbook:
        """Get the workbook object"""
        return self.workbook

    # ========================================================================
    # DOMAIN-SPECIFIC EXPORT METHODS
    # ========================================================================

    @staticmethod
    def export_voyages(queryset: QuerySet, filepath: str) -> str:
        """
        Export voyages to Excel

        Args:
            queryset: QuerySet of Voyage objects
            filepath: Path to save file

        Returns:
            str: Filepath where file was saved
        """
        service = ExcelExportService(title="Voyages Export")

        headers = [
            'Voyage Number', 'Vessel Name', 'IMO Number', 'Charter Type',
            'Charter Party', 'Load Port', 'Discharge Port',
            'Laycan Start', 'Laycan End', 'Ship Owner', 'Demurrage Rate',
            'Laytime Allowed', 'Currency', 'Assignment Status',
            'Assigned Analyst', 'Created At'
        ]

        data = []
        for voyage in queryset:
            data.append({
                'Voyage Number': voyage.voyage_number,
                'Vessel Name': voyage.vessel_name,
                'IMO Number': voyage.imo_number or 'N/A',
                'Charter Type': voyage.get_charter_type_display(),
                'Charter Party': voyage.charter_party,
                'Load Port': voyage.load_port,
                'Discharge Port': voyage.discharge_port,
                'Laycan Start': voyage.laycan_start.strftime('%Y-%m-%d'),
                'Laycan End': voyage.laycan_end.strftime('%Y-%m-%d'),
                'Ship Owner': voyage.ship_owner.name,
                'Demurrage Rate': f'${voyage.demurrage_rate:,.2f}',
                'Laytime Allowed': f'{voyage.laytime_allowed} days',
                'Currency': voyage.currency,
                'Assignment Status': voyage.get_assignment_status_display(),
                'Assigned Analyst': voyage.assigned_analyst.get_full_name() if voyage.assigned_analyst else 'Unassigned',
                'Created At': voyage.created_at.strftime('%Y-%m-%d %H:%M'),
            })

        service.create_worksheet('Voyages', headers, data)

        # Add summary
        summary = {
            'Total Voyages': queryset.count(),
            'Assigned': queryset.filter(assignment_status='ASSIGNED').count(),
            'Unassigned': queryset.filter(assignment_status='UNASSIGNED').count(),
            'Completed': queryset.filter(assignment_status='COMPLETED').count(),
        }
        service.add_summary_sheet(summary)

        return service.save(filepath)

    @staticmethod
    def export_claims(queryset: QuerySet, filepath: str) -> str:
        """
        Export claims to Excel

        Args:
            queryset: QuerySet of Claim objects
            filepath: Path to save file

        Returns:
            str: Filepath where file was saved
        """
        service = ExcelExportService(title="Claims Export")

        headers = [
            'Claim Number', 'Claim Type', 'Status', 'Payment Status',
            'Voyage Number', 'Vessel Name', 'Ship Owner',
            'Claim Amount', 'Paid Amount', 'Outstanding Amount',
            'Currency', 'Laytime Used', 'Demurrage Days',
            'Claim Deadline', 'Time Barred', 'Assigned To',
            'Created At', 'Submitted At'
        ]

        data = []
        for claim in queryset:
            data.append({
                'Claim Number': claim.claim_number,
                'Claim Type': claim.get_claim_type_display(),
                'Status': claim.get_status_display(),
                'Payment Status': claim.get_payment_status_display(),
                'Voyage Number': claim.voyage.voyage_number,
                'Vessel Name': claim.voyage.vessel_name,
                'Ship Owner': claim.ship_owner.name,
                'Claim Amount': f'${claim.claim_amount:,.2f}',
                'Paid Amount': f'${claim.paid_amount:,.2f}',
                'Outstanding Amount': f'${claim.outstanding_amount:,.2f}',
                'Currency': claim.currency,
                'Laytime Used': f'{claim.laytime_used} days' if claim.laytime_used else 'N/A',
                'Demurrage Days': f'{claim.demurrage_days} days',
                'Claim Deadline': claim.claim_deadline.strftime('%Y-%m-%d') if claim.claim_deadline else 'N/A',
                'Time Barred': 'Yes' if claim.is_time_barred else 'No',
                'Assigned To': claim.assigned_to.get_full_name() if claim.assigned_to else 'Unassigned',
                'Created At': claim.created_at.strftime('%Y-%m-%d %H:%M'),
                'Submitted At': claim.submitted_at.strftime('%Y-%m-%d %H:%M') if claim.submitted_at else 'Not Submitted',
            })

        service.create_worksheet('Claims', headers, data)

        # Add summary
        from decimal import Decimal
        total_amount = sum(c.claim_amount for c in queryset) or Decimal('0')
        paid_amount = sum(c.paid_amount for c in queryset) or Decimal('0')
        outstanding = sum(c.outstanding_amount for c in queryset) or Decimal('0')

        summary = {
            'Total Claims': queryset.count(),
            'Total Claim Amount': f'${total_amount:,.2f}',
            'Total Paid': f'${paid_amount:,.2f}',
            'Total Outstanding': f'${outstanding:,.2f}',
            'Draft Claims': queryset.filter(status='DRAFT').count(),
            'Submitted Claims': queryset.filter(status='SUBMITTED').count(),
            'Settled Claims': queryset.filter(status='SETTLED').count(),
            'Time-Barred Claims': queryset.filter(is_time_barred=True).count(),
        }
        service.add_summary_sheet(summary)

        return service.save(filepath)

    @staticmethod
    def export_ship_owners(queryset: QuerySet, filepath: str) -> str:
        """
        Export ship owners to Excel

        Args:
            queryset: QuerySet of ShipOwner objects
            filepath: Path to save file

        Returns:
            str: Filepath where file was saved
        """
        service = ExcelExportService(title="Ship Owners Export")

        headers = [
            'Name', 'Code', 'Contact Email', 'Contact Phone',
            'Address', 'Active', 'Voyages Count', 'Claims Count',
            'Created At'
        ]

        data = []
        for owner in queryset:
            data.append({
                'Name': owner.name,
                'Code': owner.code,
                'Contact Email': owner.contact_email or 'N/A',
                'Contact Phone': owner.contact_phone or 'N/A',
                'Address': owner.address or 'N/A',
                'Active': 'Yes' if owner.is_active else 'No',
                'Voyages Count': owner.voyages.count(),
                'Claims Count': owner.claims.count(),
                'Created At': owner.created_at.strftime('%Y-%m-%d'),
            })

        service.create_worksheet('Ship Owners', headers, data)

        # Add summary
        summary = {
            'Total Ship Owners': queryset.count(),
            'Active': queryset.filter(is_active=True).count(),
            'Inactive': queryset.filter(is_active=False).count(),
        }
        service.add_summary_sheet(summary)

        return service.save(filepath)

    @staticmethod
    def export_port_activities(queryset: QuerySet, filepath: str) -> str:
        """
        Export port activities to Excel

        Args:
            queryset: QuerySet of PortActivity objects
            filepath: Path to save file

        Returns:
            str: Filepath where file was saved
        """
        service = ExcelExportService(title="Port Activities Export")

        headers = [
            'Ship Name', 'Voyage Number', 'Activity Type', 'Port Name',
            'Load Port', 'Discharge Port', 'Start DateTime', 'End DateTime',
            'Duration (hours)', 'Duration (days)', 'Date Status',
            'Cargo Quantity', 'Created At'
        ]

        data = []
        for activity in queryset:
            data.append({
                'Ship Name': activity.ship.vessel_name,
                'Voyage Number': activity.voyage.voyage_number if activity.voyage else 'N/A',
                'Activity Type': activity.activity_type.name,
                'Port Name': activity.port_name,
                'Load Port': activity.load_port or 'N/A',
                'Discharge Port': activity.discharge_port or 'N/A',
                'Start DateTime': activity.start_datetime.strftime('%Y-%m-%d %H:%M'),
                'End DateTime': activity.end_datetime.strftime('%Y-%m-%d %H:%M'),
                'Duration (hours)': f'{activity.duration_hours:.2f}',
                'Duration (days)': activity.duration_days,
                'Date Status': activity.date_status_display,
                'Cargo Quantity': f'{activity.cargo_quantity:,.2f} tons' if activity.cargo_quantity else 'N/A',
                'Created At': activity.created_at.strftime('%Y-%m-%d'),
            })

        service.create_worksheet('Port Activities', headers, data)

        # Add summary
        summary = {
            'Total Activities': queryset.count(),
            'Total Duration (hours)': f'{sum(a.duration_hours for a in queryset):,.2f}',
        }
        service.add_summary_sheet(summary)

        return service.save(filepath)
