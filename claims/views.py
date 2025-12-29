from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Count, Sum, Avg, F, Case, When
from django.db import transaction
from django.utils import timezone
from datetime import datetime
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from .models import Claim, Comment, Document, User, Voyage, ShipOwner, ClaimActivityLog
from .forms import (ClaimForm, CommentForm, DocumentForm, ClaimStatusForm,
                    UserRegistrationForm, UserProfileEditForm, AdminUserEditForm)


def log_claim_activity(claim, user, action, message, old_value='', new_value=''):
    """Helper function to create activity log entries"""
    ClaimActivityLog.objects.create(
        claim=claim,
        claim_number=claim.claim_number,
        user=user,
        action=action,
        message=message,
        old_value=old_value,
        new_value=new_value
    )


def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid username or password')

    return render(request, 'claims/login.html')


def logout_view(request):
    logout(request)
    return redirect('login')


def register_view(request):
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, 'Registration successful!')
            return redirect('dashboard')
    else:
        form = UserRegistrationForm()

    return render(request, 'claims/register.html', {'form': form})


@login_required
def dashboard(request):
    user = request.user

    # Get user's claims queryset
    my_claims_queryset = Claim.objects.filter(
        Q(created_by=user) | Q(assigned_to=user)
    ).distinct().select_related('voyage', 'ship_owner')

    # Statistics
    stats = {
        'total_claims': my_claims_queryset.count(),
        'draft_claims': my_claims_queryset.filter(status='DRAFT').count(),
        'under_review': my_claims_queryset.filter(status='UNDER_REVIEW').count(),
        'submitted': my_claims_queryset.filter(status='SUBMITTED').count(),
        'settled': my_claims_queryset.filter(status='SETTLED').count(),
        'time_barred': my_claims_queryset.filter(is_time_barred=True).count(),
        'not_sent': my_claims_queryset.filter(payment_status='NOT_SENT').count(),
    }

    # Voyage statistics for analysts
    if user.can_write():
        my_voyages = Voyage.objects.filter(assigned_analyst=user)
        unassigned_voyages = Voyage.objects.filter(assignment_status='UNASSIGNED').count()
        stats['my_voyages'] = my_voyages.count()
        stats['unassigned_voyages'] = unassigned_voyages
    else:
        stats['my_voyages'] = 0
        stats['unassigned_voyages'] = 0

    # Recent claims
    recent_claims = my_claims_queryset[:10]

    context = {
        'my_claims': recent_claims,
        'stats': stats,
    }

    return render(request, 'claims/dashboard.html', context)


@login_required
def voyage_list(request):
    """Voyage Assignment Page - Shows unassigned voyages first"""
    user = request.user

    if not user.can_write():
        messages.error(request, 'You do not have permission to access voyage assignments')
        return redirect('dashboard')

    # Get all voyages
    voyages = Voyage.objects.all().select_related('ship_owner', 'assigned_analyst').prefetch_related('claims')

    # Filter by assignment status
    status_filter = request.GET.get('status', 'UNASSIGNED')  # Default to showing unassigned
    if status_filter and status_filter != 'ALL':
        voyages = voyages.filter(assignment_status=status_filter)

    # Filter by assigned analyst
    analyst_filter = request.GET.get('analyst')
    if analyst_filter:
        if analyst_filter == 'me':
            voyages = voyages.filter(assigned_analyst=user)
        else:
            voyages = voyages.filter(assigned_analyst_id=analyst_filter)

    # Search
    search_query = request.GET.get('search')
    if search_query:
        voyages = voyages.filter(
            Q(voyage_number__icontains=search_query) |
            Q(vessel_name__icontains=search_query) |
            Q(charter_party__icontains=search_query)
        )

    # Order: Unassigned first, then by created date
    voyages = voyages.annotate(
        claims_count=Count('claims')
    ).order_by(
        Case(
            When(assignment_status='UNASSIGNED', then=0),
            When(assignment_status='ASSIGNED', then=1),
            default=2
        ),
        '-created_at'
    )

    analysts = User.objects.filter(role__in=['WRITE', 'ADMIN'])

    context = {
        'voyages': voyages,
        'analysts': analysts,
        'status_filter': status_filter,
        'analyst_filter': analyst_filter,
    }

    return render(request, 'claims/voyage_list.html', context)


@login_required
def voyage_detail(request, pk):
    """Voyage detail page showing all claims"""
    voyage = get_object_or_404(Voyage.objects.select_related('ship_owner', 'assigned_analyst'), pk=pk)
    claims = voyage.claims.all().select_related('assigned_to', 'created_by')

    # Get all analysts for reassignment dropdown
    analysts = User.objects.filter(
        role__in=['WRITE', 'TEAM_LEAD', 'ADMIN'],
        is_active=True
    ).order_by('first_name', 'last_name')

    # Get assignment history
    from .models import VoyageAssignment
    assignment_history = voyage.assignment_history.select_related(
        'assigned_to', 'assigned_by'
    ).order_by('-assigned_at')

    context = {
        'voyage': voyage,
        'claims': claims,
        'can_assign': request.user.can_write(),
        'analysts': analysts,
        'assignment_history': assignment_history,
    }

    return render(request, 'claims/voyage_detail.html', context)


@login_required
def voyage_assign(request, pk):
    """Assign voyage to analyst (self-assignment)"""
    if not request.user.can_write():
        messages.error(request, 'You do not have permission to assign voyages')
        return redirect('voyage_list')

    voyage = get_object_or_404(Voyage, pk=pk)

    if request.method == 'POST':
        # Assign to self - wrapped in transaction for data consistency
        with transaction.atomic():
            voyage.assign_to(analyst=request.user, assigned_by=request.user)
        messages.success(request, f'Voyage {voyage.voyage_number} assigned to you. All claims have been auto-assigned.')
        return redirect('voyage_detail', pk=pk)

    return redirect('voyage_detail', pk=pk)


@login_required
def voyage_assign_to(request, pk):
    """Team Lead assignment - assign voyage to specific analyst"""
    if not request.user.can_assign_voyages():
        messages.error(request, 'You do not have permission to assign voyages to others')
        return redirect('voyage_list')

    voyage = get_object_or_404(Voyage, pk=pk)

    if request.method == 'POST':
        analyst_id = request.POST.get('analyst_id')
        if not analyst_id:
            messages.error(request, 'Please select an analyst')
            return redirect('voyage_list')

        analyst = get_object_or_404(User, pk=analyst_id)

        # Check if analyst can handle assignments
        if not analyst.can_write():
            messages.error(request, f'{analyst.get_full_name()} does not have permission to handle claims')
            return redirect('voyage_list')

        # Wrapped in transaction for data consistency
        with transaction.atomic():
            voyage.assign_to(analyst=analyst, assigned_by=request.user)
        messages.success(request, f'Voyage {voyage.voyage_number} assigned to {analyst.get_full_name()}. All claims have been auto-assigned.')
        return redirect('voyage_list')

    return redirect('voyage_list')


@login_required
def voyage_reassign(request, pk):
    """Reassign voyage and all its claims to a different analyst"""
    voyage = get_object_or_404(Voyage, pk=pk)

    # Permission check: only team leads, admins, or the currently assigned analyst can reassign
    can_reassign = (
        request.user.can_assign_voyages() or
        voyage.assigned_analyst == request.user
    )

    if not can_reassign:
        messages.error(request, 'You do not have permission to reassign this voyage')
        return redirect('voyage_detail', pk=pk)

    if request.method == 'POST':
        new_analyst_id = request.POST.get('new_analyst_id')
        reassignment_reason = request.POST.get('reassignment_reason', '')

        if not new_analyst_id:
            messages.error(request, 'Please select a new analyst')
            return redirect('voyage_detail', pk=pk)

        new_analyst = get_object_or_404(User, pk=new_analyst_id)

        if not new_analyst.can_write():
            messages.error(request, f'{new_analyst.get_full_name()} does not have permission to handle claims')
            return redirect('voyage_detail', pk=pk)

        old_analyst = voyage.assigned_analyst

        # Reassign voyage and all claims, add comments - all in one transaction
        with transaction.atomic():
            # Reassign voyage and all claims
            voyage.assign_to(analyst=new_analyst, assigned_by=request.user)

            # Update the assignment history record with reassignment reason
            if reassignment_reason:
                from .models import VoyageAssignment
                latest_assignment = VoyageAssignment.objects.filter(
                    voyage=voyage,
                    is_active=True
                ).first()
                if latest_assignment:
                    latest_assignment.reassignment_reason = reassignment_reason
                    latest_assignment.save()

                # Add a comment to all claims about the reassignment
                for claim in voyage.claims.all():
                    Comment.objects.create(
                        claim=claim,
                        user=request.user,
                        content=f'Reassigned from {old_analyst.get_full_name() if old_analyst else "Unassigned"} to {new_analyst.get_full_name()}. Reason: {reassignment_reason}'
                    )

        messages.success(request, f'Voyage {voyage.voyage_number} and all claims reassigned to {new_analyst.get_full_name()}')
        return redirect('voyage_detail', pk=pk)

    return redirect('voyage_detail', pk=pk)


@login_required
def analytics_dashboard(request):
    """Analytics and reporting dashboard"""
    user = request.user

    # Get all claims
    claims = Claim.objects.all().select_related('voyage', 'ship_owner', 'assigned_to')

    # Apply filters
    owner_filter = request.GET.get('owner')
    if owner_filter:
        claims = claims.filter(ship_owner_id=owner_filter)

    analyst_filter = request.GET.get('analyst')
    if analyst_filter:
        if analyst_filter == 'me':
            claims = claims.filter(assigned_to=user)
        else:
            claims = claims.filter(assigned_to_id=analyst_filter)

    payment_filter = request.GET.get('payment')
    if payment_filter:
        claims = claims.filter(payment_status=payment_filter)

    timebar_filter = request.GET.get('timebar')
    if timebar_filter == 'yes':
        claims = claims.filter(is_time_barred=True)
    elif timebar_filter == 'no':
        claims = claims.filter(is_time_barred=False)

    # Aggregate by Ship Owner
    owner_stats = claims.values('ship_owner__name', 'ship_owner__code').annotate(
        total_claims=Count('id'),
        total_value=Sum('claim_amount'),
        total_paid=Sum('paid_amount'),
        total_outstanding=Sum(F('claim_amount') - F('paid_amount')),
        time_barred_count=Count(Case(When(is_time_barred=True, then=1))),
    ).order_by('-total_value')

    # Aggregate by Analyst
    analyst_stats = claims.values('assigned_to__username', 'assigned_to__first_name', 'assigned_to__last_name').annotate(
        total_claims=Count('id'),
        total_value=Sum('claim_amount'),
        total_paid=Sum('paid_amount'),
        total_outstanding=Sum(F('claim_amount') - F('paid_amount')),
        time_barred_count=Count(Case(When(is_time_barred=True, then=1))),
        avg_value=Avg('claim_amount'),
    ).order_by('-total_value')

    # Add analyst_name for easier template access
    for stat in analyst_stats:
        first_name = stat.get('assigned_to__first_name', '')
        last_name = stat.get('assigned_to__last_name', '')
        username = stat.get('assigned_to__username', '')
        if first_name and last_name:
            stat['analyst_name'] = f"{first_name} {last_name}"
        elif first_name:
            stat['analyst_name'] = first_name
        else:
            stat['analyst_name'] = username or 'Unassigned'

    # Aggregate by Vessel
    vessel_stats = claims.values('voyage__vessel_name').annotate(
        total_claims=Count('id'),
        total_value=Sum('claim_amount'),
        total_paid=Sum('paid_amount'),
        total_outstanding=Sum(F('claim_amount') - F('paid_amount')),
    ).order_by('-total_value')[:20]  # Top 20 vessels

    # Overall statistics (calculate first, needed for percentages)
    overall_stats = claims.aggregate(
        total_claims=Count('id'),
        total_value=Sum('claim_amount'),
        total_paid=Sum('paid_amount'),
        total_outstanding=Sum(F('claim_amount') - F('paid_amount')),
        time_barred_count=Count(Case(When(is_time_barred=True, then=1))),
    )

    # Payment status breakdown
    payment_breakdown = claims.values('payment_status').annotate(
        count=Count('id'),
        total_value=Sum('claim_amount')
    ).order_by('payment_status')

    # Add status display and percentage to payment breakdown
    status_dict = dict(Claim.PAYMENT_STATUS_CHOICES)
    for stat in payment_breakdown:
        stat['status_display'] = status_dict.get(stat['payment_status'], stat['payment_status'])
        if overall_stats['total_value']:
            stat['percentage'] = (stat['total_value'] / overall_stats['total_value']) * 100
        else:
            stat['percentage'] = 0

    # Get filter options
    ship_owners = ShipOwner.objects.all().order_by('name')
    analysts = User.objects.filter(role__in=['WRITE', 'ADMIN']).order_by('first_name')

    context = {
        'owner_stats': owner_stats,
        'analyst_stats': analyst_stats,
        'vessel_stats': vessel_stats,
        'payment_breakdown': payment_breakdown,
        'overall_stats': overall_stats,
        'ship_owners': ship_owners,
        'analysts': analysts,
        'filters': {
            'owner': owner_filter,
            'analyst': analyst_filter,
            'payment': payment_filter,
            'timebar': timebar_filter,
        }
    }

    return render(request, 'claims/analytics.html', context)


@login_required
def export_payment_breakdown(request):
    """Export payment status breakdown to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    user = request.user

    # Get filtered claims (same logic as analytics_dashboard)
    claims = Claim.objects.all().select_related('voyage', 'ship_owner', 'assigned_to')

    # Apply same filters
    owner_filter = request.GET.get('owner')
    if owner_filter:
        claims = claims.filter(ship_owner_id=owner_filter)

    analyst_filter = request.GET.get('analyst')
    if analyst_filter:
        if analyst_filter == 'me':
            claims = claims.filter(assigned_to=user)
        else:
            claims = claims.filter(assigned_to_id=analyst_filter)

    payment_filter = request.GET.get('payment')
    if payment_filter:
        claims = claims.filter(payment_status=payment_filter)

    timebar_filter = request.GET.get('timebar')
    if timebar_filter == 'yes':
        claims = claims.filter(is_time_barred=True)
    elif timebar_filter == 'no':
        claims = claims.filter(is_time_barred=False)

    # Get payment breakdown
    payment_breakdown = claims.values('payment_status').annotate(
        count=Count('id'),
        total_value=Sum('claim_amount')
    ).order_by('payment_status')

    # Calculate total for percentages
    total_value = claims.aggregate(total=Sum('claim_amount'))['total'] or 0

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Payment Status Breakdown"

    # Headers
    headers = ['Payment Status', 'Count', 'Total Value (USD)', '% of Total']
    header_fill = PatternFill(start_color='0dcaf0', end_color='0dcaf0', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font

    # Data rows
    status_dict = dict(Claim.PAYMENT_STATUS_CHOICES)
    for row_num, stat in enumerate(payment_breakdown, 2):
        ws.cell(row=row_num, column=1, value=status_dict.get(stat['payment_status'], stat['payment_status']))
        ws.cell(row=row_num, column=2, value=stat['count'])
        ws.cell(row=row_num, column=3, value=float(stat['total_value'] or 0))
        percentage = (float(stat['total_value'] or 0) / total_value * 100) if total_value > 0 else 0
        ws.cell(row=row_num, column=4, value=f"{percentage:.1f}%")

    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=payment_breakdown_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    wb.save(response)
    return response


@login_required
def export_owner_stats(request):
    """Export ship owner statistics to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    user = request.user

    # Get filtered claims (same logic as analytics_dashboard)
    claims = Claim.objects.all().select_related('voyage', 'ship_owner', 'assigned_to')

    # Apply same filters
    owner_filter = request.GET.get('owner')
    if owner_filter:
        claims = claims.filter(ship_owner_id=owner_filter)

    analyst_filter = request.GET.get('analyst')
    if analyst_filter:
        if analyst_filter == 'me':
            claims = claims.filter(assigned_to=user)
        else:
            claims = claims.filter(assigned_to_id=analyst_filter)

    payment_filter = request.GET.get('payment')
    if payment_filter:
        claims = claims.filter(payment_status=payment_filter)

    timebar_filter = request.GET.get('timebar')
    if timebar_filter == 'yes':
        claims = claims.filter(is_time_barred=True)
    elif timebar_filter == 'no':
        claims = claims.filter(is_time_barred=False)

    # Aggregate by Ship Owner
    owner_stats = claims.values('ship_owner__name', 'ship_owner__code').annotate(
        total_claims=Count('id'),
        total_value=Sum('claim_amount'),
        total_paid=Sum('paid_amount'),
        total_outstanding=Sum(F('claim_amount') - F('paid_amount')),
        time_barred_count=Count(Case(When(is_time_barred=True, then=1))),
    ).order_by('-total_value')

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Ship Owner Statistics"

    # Headers
    headers = ['Ship Owner', 'Owner Code', 'Total Claims', 'Total Value (USD)',
               'Total Paid (USD)', 'Outstanding (USD)', 'Time-barred']
    header_fill = PatternFill(start_color='0d6efd', end_color='0d6efd', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font

    # Data rows
    for row_num, stat in enumerate(owner_stats, 2):
        ws.cell(row=row_num, column=1, value=stat['ship_owner__name'])
        ws.cell(row=row_num, column=2, value=stat['ship_owner__code'] or '')
        ws.cell(row=row_num, column=3, value=stat['total_claims'])
        ws.cell(row=row_num, column=4, value=float(stat['total_value'] or 0))
        ws.cell(row=row_num, column=5, value=float(stat['total_paid'] or 0))
        ws.cell(row=row_num, column=6, value=float(stat['total_outstanding'] or 0))
        ws.cell(row=row_num, column=7, value=stat['time_barred_count'])

    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=owner_stats_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    wb.save(response)
    return response


@login_required
def claim_list(request):
    user = request.user
    claims = Claim.objects.all().select_related('voyage', 'ship_owner', 'assigned_to', 'created_by')

    # Filter by status
    status_filter = request.GET.get('status')
    if status_filter:
        claims = claims.filter(status=status_filter)

    # Filter by payment status
    payment_filter = request.GET.get('payment')
    if payment_filter:
        claims = claims.filter(payment_status=payment_filter)

    # Filter by type
    type_filter = request.GET.get('claim_type')
    if type_filter:
        claims = claims.filter(claim_type=type_filter)

    # Filter by timebar
    timebar_filter = request.GET.get('timebar')
    if timebar_filter == 'yes':
        claims = claims.filter(is_time_barred=True)

    # Search
    search_query = request.GET.get('search')
    if search_query:
        claims = claims.filter(
            Q(claim_number__icontains=search_query) |
            Q(voyage__vessel_name__icontains=search_query) |
            Q(voyage__voyage_number__icontains=search_query) |
            Q(ship_owner__name__icontains=search_query)
        )

    # Apply permissions
    if not (user.can_write() or user.is_admin_role()):
        claims = claims.filter(Q(created_by=user) | Q(assigned_to=user)).distinct()

    context = {
        'claims': claims,
        'status_choices': Claim.STATUS_CHOICES,
        'payment_choices': Claim.PAYMENT_STATUS_CHOICES,
        'type_choices': Claim.CLAIM_TYPE_CHOICES,
    }

    return render(request, 'claims/claim_list.html', context)


@login_required
def claim_detail(request, pk):
    claim = get_object_or_404(
        Claim.objects.select_related('voyage', 'ship_owner', 'assigned_to', 'created_by'),
        pk=pk
    )
    user = request.user

    # Check permissions
    if not (user.is_admin_role() or user.can_write() or claim.created_by == user or claim.assigned_to == user):
        messages.error(request, 'You do not have permission to view this claim')
        return redirect('claim_list')

    comments = claim.comments.all().select_related('user')
    documents = claim.documents.all().select_related('uploaded_by')
    activity_logs = claim.activity_logs.all().select_related('user').order_by('-created_at')

    context = {
        'claim': claim,
        'comments': comments,
        'documents': documents,
        'activity_logs': activity_logs,
        'comment_form': CommentForm(),
        'document_form': DocumentForm(),
        'status_form': ClaimStatusForm(instance=claim),
        'can_edit': claim.can_edit(user),
        'can_delete': claim.can_delete(user),
    }

    return render(request, 'claims/claim_detail.html', context)


@login_required
def claim_create(request):
    if not request.user.can_write():
        messages.error(request, 'You do not have permission to create claims')
        return redirect('claim_list')

    if request.method == 'POST':
        form = ClaimForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                claim = form.save(commit=False)
                claim.created_by = request.user
                claim.ship_owner = claim.voyage.ship_owner  # Auto-set from voyage
                claim.save()

                # Log claim creation
                log_claim_activity(
                    claim, request.user, 'CREATED',
                    f'Claim created for {claim.voyage.vessel_name} ({claim.get_claim_type_display()})',
                    new_value=f'{claim.claim_amount} {claim.currency}'
                )
            messages.success(request, f'Claim {claim.claim_number} created successfully')
            return redirect('claim_detail', pk=claim.pk)
    else:
        form = ClaimForm()

    return render(request, 'claims/claim_form.html', {'form': form, 'title': 'Create Claim'})


@login_required
def claim_update(request, pk):
    claim = get_object_or_404(Claim, pk=pk)

    if not claim.can_edit(request.user):
        messages.error(request, 'You do not have permission to edit this claim')
        return redirect('claim_detail', pk=pk)

    if request.method == 'POST':
        # Get old values before update
        old_claim_amount = claim.claim_amount
        old_paid_amount = claim.paid_amount
        old_deadline = claim.claim_deadline
        old_assigned_to = claim.assigned_to

        form = ClaimForm(request.POST, instance=claim)
        if form.is_valid():
            with transaction.atomic():
                claim = form.save(commit=False)
                claim.ship_owner = claim.voyage.ship_owner  # Ensure consistency
                claim.save()

                # Log significant changes
                if old_claim_amount != claim.claim_amount:
                    log_claim_activity(claim, request.user, 'AMOUNT_CHANGED', 'Claim amount changed',
                        old_value=f'{old_claim_amount} {claim.currency}',
                        new_value=f'{claim.claim_amount} {claim.currency}')

                if old_paid_amount != claim.paid_amount:
                    log_claim_activity(claim, request.user, 'PAID_AMOUNT_CHANGED', 'Paid amount updated',
                        old_value=f'{old_paid_amount} {claim.currency}',
                        new_value=f'{claim.paid_amount} {claim.currency}')

                if old_deadline != claim.claim_deadline:
                    log_claim_activity(claim, request.user, 'DEADLINE_CHANGED', 'Claim deadline changed',
                        old_value=str(old_deadline) if old_deadline else 'None',
                        new_value=str(claim.claim_deadline) if claim.claim_deadline else 'None')

                if old_assigned_to != claim.assigned_to:
                    action = 'REASSIGNED' if old_assigned_to else 'ASSIGNED'
                    log_claim_activity(claim, request.user, action,
                        f'Claim {"reassigned" if old_assigned_to else "assigned"}',
                        old_value=old_assigned_to.get_full_name() if old_assigned_to else 'Unassigned',
                        new_value=claim.assigned_to.get_full_name() if claim.assigned_to else 'Unassigned')

            messages.success(request, 'Claim updated successfully')
            return redirect('claim_detail', pk=pk)
    else:
        form = ClaimForm(instance=claim)

    return render(request, 'claims/claim_form.html', {'form': form, 'title': 'Edit Claim', 'claim': claim})


@login_required
def claim_delete(request, pk):
    claim = get_object_or_404(Claim, pk=pk)

    if not claim.can_delete(request.user):
        messages.error(request, 'You do not have permission to delete this claim')
        return redirect('claim_detail', pk=pk)

    if request.method == 'POST':
        with transaction.atomic():
            # Log deletion before deleting the claim
            log_claim_activity(claim, request.user, 'DELETED',
                f'Claim deleted',
                old_value=f'{claim.get_claim_type_display()} - {claim.claim_amount} {claim.currency}')
            claim.delete()
        messages.success(request, 'Claim deleted successfully')
        return redirect('claim_list')

    return render(request, 'claims/claim_confirm_delete.html', {'claim': claim})


@login_required
def claim_status_update(request, pk):
    claim = get_object_or_404(Claim, pk=pk)

    if not claim.can_edit(request.user):
        messages.error(request, 'You do not have permission to update this claim status')
        return redirect('claim_detail', pk=pk)

    if request.method == 'POST':
        # Get old values before update
        old_status = claim.status
        old_payment_status = claim.payment_status

        form = ClaimStatusForm(request.POST, instance=claim)
        if form.is_valid():
            # Wrapped in transaction to ensure status + timestamp updates are atomic
            with transaction.atomic():
                claim = form.save(commit=False)

                # Update timestamps based on status changes
                if claim.status == 'SUBMITTED' and not claim.submitted_at:
                    claim.submitted_at = timezone.now()
                elif claim.status == 'SETTLED' and not claim.settled_at:
                    claim.settled_at = timezone.now()

                # Update payment timestamps
                if claim.payment_status == 'SENT' and not claim.sent_to_owner_at:
                    claim.sent_to_owner_at = timezone.now()
                elif claim.payment_status in ['PAID', 'PARTIALLY_PAID'] and not claim.paid_at:
                    claim.paid_at = timezone.now()

                claim.save()

                # Log status changes
                if old_status != claim.status:
                    ClaimActivityLog.objects.create(
                        claim=claim,
                        user=request.user,
                        action='STATUS_CHANGED',
                        message=f'Claim status changed',
                        old_value=dict(Claim.STATUS_CHOICES).get(old_status, old_status),
                        new_value=dict(Claim.STATUS_CHOICES).get(claim.status, claim.status)
                    )

                if old_payment_status != claim.payment_status:
                    ClaimActivityLog.objects.create(
                        claim=claim,
                        user=request.user,
                        action='PAYMENT_STATUS_CHANGED',
                        message=f'Payment status changed',
                        old_value=dict(Claim.PAYMENT_STATUS_CHOICES).get(old_payment_status, old_payment_status),
                        new_value=dict(Claim.PAYMENT_STATUS_CHOICES).get(claim.payment_status, claim.payment_status)
                    )

                    # Log if claim becomes time-barred
                    if claim.payment_status == 'TIMEBAR' and old_payment_status != 'TIMEBAR':
                        ClaimActivityLog.objects.create(
                            claim=claim,
                            user=request.user,
                            action='TIME_BARRED',
                            message=f'Claim marked as time-barred',
                            new_value=str(claim.time_bar_date) if claim.time_bar_date else timezone.now().date().isoformat()
                        )

            messages.success(request, 'Claim status updated successfully')
            return redirect('claim_detail', pk=pk)

    return redirect('claim_detail', pk=pk)


@login_required
def add_comment(request, claim_pk):
    claim = get_object_or_404(Claim, pk=claim_pk)

    if request.method == 'POST':
        form = CommentForm(request.POST)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.claim = claim
            comment.user = request.user
            comment.save()
            messages.success(request, 'Comment added successfully')

    return redirect('claim_detail', pk=claim_pk)


@login_required
def add_document(request, claim_pk):
    claim = get_object_or_404(Claim, pk=claim_pk)

    if not claim.can_edit(request.user):
        messages.error(request, 'You do not have permission to add documents to this claim')
        return redirect('claim_detail', pk=claim_pk)

    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES)
        if form.is_valid():
            document = form.save(commit=False)
            document.claim = claim
            document.uploaded_by = request.user
            document.save()
            messages.success(request, 'Document uploaded successfully')

    return redirect('claim_detail', pk=claim_pk)


@login_required
def export_claims(request):
    if not request.user.can_export():
        messages.error(request, 'You do not have permission to export claims')
        return redirect('claim_list')

    user = request.user
    claims = Claim.objects.all().select_related('voyage', 'ship_owner', 'assigned_to', 'created_by')

    # Apply same filters as list view
    if not (user.can_write() or user.is_admin_role()):
        claims = claims.filter(Q(created_by=user) | Q(assigned_to=user)).distinct()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Claims"

    # Headers
    headers = [
        'Claim Number', 'RADAR ID', 'Voyage Number', 'Vessel Name', 'Ship Owner',
        'Type', 'Cost Type', 'Status', 'Payment Status', 'Claim Amount', 'Paid Amount',
        'Outstanding', 'Currency', 'Time Barred', 'Claim Deadline', 'Assigned To',
        'Created At'
    ]

    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font

    # Data rows
    for row_num, claim in enumerate(claims, 2):
        ws.cell(row=row_num, column=1, value=claim.claim_number)
        ws.cell(row=row_num, column=2, value=claim.radar_claim_id or '')
        ws.cell(row=row_num, column=3, value=claim.voyage.voyage_number)
        ws.cell(row=row_num, column=4, value=claim.voyage.vessel_name)
        ws.cell(row=row_num, column=5, value=claim.ship_owner.name)
        ws.cell(row=row_num, column=6, value=claim.get_claim_type_display())
        ws.cell(row=row_num, column=7, value=claim.get_cost_type_display() if claim.cost_type else '')
        ws.cell(row=row_num, column=8, value=claim.get_status_display())
        ws.cell(row=row_num, column=9, value=claim.get_payment_status_display())
        ws.cell(row=row_num, column=10, value=float(claim.claim_amount))
        ws.cell(row=row_num, column=11, value=float(claim.paid_amount))
        ws.cell(row=row_num, column=12, value=float(claim.outstanding_amount))
        ws.cell(row=row_num, column=13, value=claim.currency)
        ws.cell(row=row_num, column=14, value='Yes' if claim.is_time_barred else 'No')
        ws.cell(row=row_num, column=15, value=claim.claim_deadline.strftime('%Y-%m-%d') if claim.claim_deadline else '')
        ws.cell(row=row_num, column=16, value=claim.assigned_to.get_full_name() if claim.assigned_to else '')
        ws.cell(row=row_num, column=17, value=claim.created_at.strftime('%Y-%m-%d %H:%M'))

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=claims_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    wb.save(response)
    return response


# ============================================================================
# USER PROFILE VIEWS
# ============================================================================

@login_required
def user_profile(request, user_id):
    """View user profile - anyone can view any profile"""
    profile_user = get_object_or_404(User, pk=user_id)
    
    # Get statistics
    assigned_voyages_count = profile_user.get_assigned_voyages_count()
    closed_claims_count = profile_user.get_closed_claims_count()
    total_claims = profile_user.assigned_claims.count()
    
    # Recent activity
    recent_claims = profile_user.assigned_claims.select_related(
        'voyage', 'ship_owner'
    ).order_by('-updated_at')[:5]
    
    recent_voyages = profile_user.assigned_voyages.select_related(
        'ship_owner'
    ).order_by('-assigned_at')[:5]
    
    context = {
        'profile_user': profile_user,
        'assigned_voyages_count': assigned_voyages_count,
        'closed_claims_count': closed_claims_count,
        'total_claims': total_claims,
        'recent_claims': recent_claims,
        'recent_voyages': recent_voyages,
    }
    
    return render(request, 'claims/user_profile.html', context)


@login_required
def user_profile_edit(request, user_id):
    """Edit user profile - users can edit their own, admins can edit anyone's"""
    profile_user = get_object_or_404(User, pk=user_id)
    
    # Permission check
    if not (request.user == profile_user or request.user.is_admin_role()):
        messages.error(request, "You don't have permission to edit this profile.")
        return redirect('user_profile', user_id=user_id)
    
    if request.method == 'POST':
        # Use appropriate form based on user role
        if request.user.is_admin_role():
            form = AdminUserEditForm(request.POST, request.FILES, instance=profile_user)
        else:
            form = UserProfileEditForm(request.POST, request.FILES, instance=profile_user)
        
        if form.is_valid():
            form.save()
            messages.success(request, 'Profile updated successfully!')
            return redirect('user_profile', user_id=user_id)
    else:
        # Use appropriate form based on user role
        if request.user.is_admin_role():
            form = AdminUserEditForm(instance=profile_user)
        else:
            form = UserProfileEditForm(instance=profile_user)
    
    context = {
        'form': form,
        'profile_user': profile_user,
    }
    
    return render(request, 'claims/user_profile_edit.html', context)


@login_required
def user_directory(request):
    """User directory - search and browse all users"""
    # Check if user has permission to view user directory
    if not (request.user.has_perm('claims.view_user_directory') or request.user.is_staff):
        messages.error(request, 'You do not have permission to view the user directory')
        return redirect('dashboard')

    search_query = request.GET.get('search', '')
    role_filter = request.GET.get('role', '')
    department_filter = request.GET.get('department', '')
    view_mode = request.GET.get('view', 'cards')  # cards or table
    sort_by = request.GET.get('sort', 'name')  # name, created, voyages, claims
    sort_order = request.GET.get('order', 'asc')  # asc or desc

    users = User.objects.filter(is_active=True)

    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query)
        )

    if role_filter:
        users = users.filter(role=role_filter)

    if department_filter:
        users = users.filter(department__icontains=department_filter)

    # Annotate with statistics
    users = users.annotate(
        voyages_count=Count('assigned_voyages'),
        claims_count=Count('assigned_claims')
    )

    # Apply sorting
    order_prefix = '-' if sort_order == 'desc' else ''
    if sort_by == 'name':
        users = users.order_by(f'{order_prefix}last_name', f'{order_prefix}first_name')
    elif sort_by == 'created':
        users = users.order_by(f'{order_prefix}date_joined')
    elif sort_by == 'voyages':
        users = users.order_by(f'{order_prefix}voyages_count')
    elif sort_by == 'claims':
        users = users.order_by(f'{order_prefix}claims_count')
    else:
        users = users.order_by('last_name', 'first_name')

    # Get unique departments for filter
    departments = User.objects.filter(
        is_active=True, department__isnull=False
    ).exclude(department='').values_list('department', flat=True).distinct()

    context = {
        'users': users,
        'role_choices': User.ROLE_CHOICES,
        'departments': departments,
        'search_query': search_query,
        'role_filter': role_filter,
        'department_filter': department_filter,
        'view_mode': view_mode,
        'sort_by': sort_by,
        'sort_order': sort_order,
    }

    return render(request, 'claims/user_directory.html', context)


@login_required
def export_users(request):
    """Export users to Excel"""
    if not (request.user.has_perm('claims.export_users') or request.user.is_staff):
        messages.error(request, 'You do not have permission to export users')
        return redirect('user_directory')

    # Get all active users with statistics
    users = User.objects.filter(is_active=True).annotate(
        voyages_count=Count('assigned_voyages'),
        claims_count=Count('assigned_claims')
    ).order_by('last_name', 'first_name')

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"

    # Headers
    headers = [
        'Full Name', 'Username', 'Email', 'Role', 'Department',
        'Voyages Assigned', 'Claims Assigned', 'Dark Mode',
        'Date Joined', 'Last Login'
    ]

    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font

    # Data rows
    for row_num, user in enumerate(users, 2):
        ws.cell(row=row_num, column=1, value=user.get_full_name() or user.username)
        ws.cell(row=row_num, column=2, value=user.username)
        ws.cell(row=row_num, column=3, value=user.email or '')
        ws.cell(row=row_num, column=4, value=user.get_role_display())
        ws.cell(row=row_num, column=5, value=user.department or '')
        ws.cell(row=row_num, column=6, value=user.voyages_count)
        ws.cell(row=row_num, column=7, value=user.claims_count)
        ws.cell(row=row_num, column=8, value='Yes' if user.dark_mode else 'No')
        ws.cell(row=row_num, column=9, value=user.date_joined.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row_num, column=10, value=user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else 'Never')

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=users_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    wb.save(response)
    return response


@login_required
def toggle_dark_mode(request):
    """Toggle dark mode preference"""
    if request.method == 'POST':
        request.user.dark_mode = not request.user.dark_mode
        request.user.save()
        return JsonResponse({'dark_mode': request.user.dark_mode})
    return JsonResponse({'error': 'Invalid request'}, status=400)


# ============================================================================
# TC FLEET VIEWS (from ships app)
# ============================================================================

from ships.models import TCFleet, Ship


@login_required
def tc_fleet_list(request):
    """Display list of TC Fleet contracts with filters."""
    # Get filter parameters
    status_filter = request.GET.get('status', '')
    ship_type_filter = request.GET.get('ship_type', '')
    trade_filter = request.GET.get('trade', '')
    owner_filter = request.GET.get('owner', '')
    trader_filter = request.GET.get('trader', '')
    search_query = request.GET.get('search', '')

    # Base queryset
    contracts = TCFleet.objects.all().select_related('created_by')

    # Apply filters
    if status_filter:
        contracts = contracts.filter(delivery_status=status_filter)
    if ship_type_filter:
        contracts = contracts.filter(ship_type=ship_type_filter)
    if trade_filter:
        contracts = contracts.filter(trade=trade_filter)
    if owner_filter:
        contracts = contracts.filter(owner_name=owner_filter)
    if trader_filter:
        contracts = contracts.filter(trader_name=trader_filter)
    if search_query:
        contracts = contracts.filter(
            Q(ship_name__icontains=search_query) |
            Q(imo_number__icontains=search_query) |
            Q(radar_deal_number__icontains=search_query) |
            Q(owner_name__icontains=search_query)
        )

    # Get unique values for filter dropdowns
    ship_types = TCFleet.objects.values_list('ship_type', flat=True).distinct().order_by('ship_type')
    trades = TCFleet.objects.values_list('trade', flat=True).distinct().order_by('trade')
    owners = TCFleet.objects.values_list('owner_name', flat=True).distinct().order_by('owner_name')
    traders = TCFleet.objects.filter(trader_name__isnull=False).exclude(trader_name='').values_list('trader_name', flat=True).distinct().order_by('trader_name')

    # Calculate statistics
    total_contracts = contracts.count()
    active_contracts = contracts.filter(delivery_status='ON_TC').count()
    incoming_contracts = contracts.filter(delivery_status='INCOMING_TC').count()
    redelivered_contracts = contracts.filter(delivery_status='REDELIVERED').count()

    context = {
        'contracts': contracts,
        'ship_types': ship_types,
        'trades': trades,
        'owners': owners,
        'traders': traders,
        'status_filter': status_filter,
        'ship_type_filter': ship_type_filter,
        'trade_filter': trade_filter,
        'owner_filter': owner_filter,
        'trader_filter': trader_filter,
        'search_query': search_query,
        'total_contracts': total_contracts,
        'active_contracts': active_contracts,
        'incoming_contracts': incoming_contracts,
        'redelivered_contracts': redelivered_contracts,
    }

    return render(request, 'ships/tc_fleet_list.html', context)


@login_required
def tc_fleet_detail(request, pk):
    """Display detailed view of a single TC Fleet contract."""
    contract = get_object_or_404(TCFleet, pk=pk)

    # Get all contracts for the same ship (IMO)
    contract_history = TCFleet.objects.filter(
        imo_number=contract.imo_number
    ).exclude(pk=pk).order_by('-delivery_date')

    # Get ship master data if exists
    ship_master = contract.get_ship_master_data()

    context = {
        'contract': contract,
        'contract_history': contract_history,
        'ship_master': ship_master,
    }

    return render(request, 'ships/tc_fleet_detail.html', context)


@login_required
def tc_fleet_by_ship(request, imo_number):
    """Display all TC Fleet contracts for a specific ship (by IMO number)."""
    # Get all contracts for this IMO
    contracts = TCFleet.objects.filter(
        imo_number=imo_number
    ).order_by('-delivery_date')

    if not contracts.exists():
        return redirect('tc_fleet_list')

    # Get ship master data if exists
    ship_master = None
    try:
        ship_master = Ship.objects.get(imo_number=imo_number)
    except Ship.DoesNotExist:
        pass

    # Get the most recent contract for ship name
    latest_contract = contracts.first()

    # Calculate statistics for this ship
    total_contracts = contracts.count()
    active_contracts = contracts.filter(delivery_status='ON_TC').count()
    redelivered_contracts = contracts.filter(delivery_status='REDELIVERED').count()

    context = {
        'contracts': contracts,
        'imo_number': imo_number,
        'ship_name': latest_contract.ship_name,
        'ship_master': ship_master,
        'total_contracts': total_contracts,
        'active_contracts': active_contracts,
        'redelivered_contracts': redelivered_contracts,
    }

    return render(request, 'ships/tc_fleet_by_ship.html', context)


@login_required
def tc_fleet_export(request):
    """Export TC Fleet contracts to Excel."""
    # Check if user has export permission
    if not request.user.can_export():
        return redirect('tc_fleet_list')

    # Get filter parameters
    status_filter = request.GET.get('status', '')
    ship_type_filter = request.GET.get('ship_type', '')
    trade_filter = request.GET.get('trade', '')
    owner_filter = request.GET.get('owner', '')
    trader_filter = request.GET.get('trader', '')
    search_query = request.GET.get('search', '')

    # Base queryset
    contracts = TCFleet.objects.all().select_related('created_by')

    # Apply filters
    if status_filter:
        contracts = contracts.filter(delivery_status=status_filter)
    if ship_type_filter:
        contracts = contracts.filter(ship_type=ship_type_filter)
    if trade_filter:
        contracts = contracts.filter(trade=trade_filter)
    if owner_filter:
        contracts = contracts.filter(owner_name=owner_filter)
    if trader_filter:
        contracts = contracts.filter(trader_name=trader_filter)
    if search_query:
        contracts = contracts.filter(
            Q(ship_name__icontains=search_query) |
            Q(imo_number__icontains=search_query) |
            Q(radar_deal_number__icontains=search_query) |
            Q(owner_name__icontains=search_query)
        )

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "TC Fleet Contracts"

    # Define headers
    headers = [
        'Ship Name', 'IMO Number', 'Ship Type', 'Delivery Status', 'Trade',
        'Owner Name', 'Owner Email', 'Technical Manager', 'Tech Manager Email',
        'Charter Length (Years)', 'TC Rate Monthly (USD)', 'RADAR Deal Number',
        'Delivery Date', 'Redelivery Date', 'TCP Date', 'Broker Name',
        'Broker Email', 'Broker Commission (%)', 'Delivery Location',
        'Redelivery Location', 'Bunkers Policy', 'Summer DWT', 'Built Year',
        'Flag', 'Next Dry-Dock Date', 'Trader Name', 'Days Remaining',
        'Contract Status', 'Total Contract Value (USD)'
    ]

    # Style header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font

    # Write data rows
    for row_num, contract in enumerate(contracts, 2):
        ws.cell(row=row_num, column=1).value = contract.ship_name
        ws.cell(row=row_num, column=2).value = contract.imo_number
        ws.cell(row=row_num, column=3).value = contract.get_ship_type_display()
        ws.cell(row=row_num, column=4).value = contract.get_delivery_status_display()
        ws.cell(row=row_num, column=5).value = contract.get_trade_display()
        ws.cell(row=row_num, column=6).value = contract.owner_name
        ws.cell(row=row_num, column=7).value = contract.owner_email
        ws.cell(row=row_num, column=8).value = contract.technical_manager
        ws.cell(row=row_num, column=9).value = contract.technical_manager_email
        ws.cell(row=row_num, column=10).value = float(contract.charter_length_years)
        ws.cell(row=row_num, column=11).value = float(contract.tc_rate_monthly)
        ws.cell(row=row_num, column=12).value = contract.radar_deal_number
        ws.cell(row=row_num, column=13).value = contract.delivery_date.strftime('%d/%m/%Y')
        ws.cell(row=row_num, column=14).value = contract.redelivery_date.strftime('%d/%m/%Y')
        ws.cell(row=row_num, column=15).value = contract.tcp_date.strftime('%d/%m/%Y')
        ws.cell(row=row_num, column=16).value = contract.broker_name
        ws.cell(row=row_num, column=17).value = contract.broker_email
        ws.cell(row=row_num, column=18).value = float(contract.broker_commission) if contract.broker_commission else None
        ws.cell(row=row_num, column=19).value = contract.delivery_location
        ws.cell(row=row_num, column=20).value = contract.redelivery_location
        ws.cell(row=row_num, column=21).value = contract.get_bunkers_policy_display() if contract.bunkers_policy else ''
        ws.cell(row=row_num, column=22).value = float(contract.summer_dwt)
        ws.cell(row=row_num, column=23).value = contract.built_year
        ws.cell(row=row_num, column=24).value = contract.flag
        ws.cell(row=row_num, column=25).value = contract.next_drydock_date.strftime('%d/%m/%Y') if contract.next_drydock_date else ''
        ws.cell(row=row_num, column=26).value = contract.trader_name
        ws.cell(row=row_num, column=27).value = contract.days_remaining
        ws.cell(row=row_num, column=28).value = contract.contract_status
        ws.cell(row=row_num, column=29).value = float(contract.total_contract_value)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=tc_fleet_contracts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    wb.save(response)
    return response


# ============================================
# SHIP SPECIFICATIONS (Q88) VIEWS
# ============================================

@login_required
def ship_specifications_list(request):
    """
    Display list of Ship Specifications (Q88) with filters.
    """
    from ships.models import ShipSpecification

    # Get filter parameters
    vessel_type_filter = request.GET.get('vessel_type', '')
    flag_filter = request.GET.get('flag', '')
    coating_filter = request.GET.get('coating', '')
    fuel_type_filter = request.GET.get('fuel_type', '')
    tc_status_filter = request.GET.get('tc_status', '')
    search_query = request.GET.get('search', '')

    # Base queryset
    specifications = ShipSpecification.objects.all().select_related('created_by')

    # Apply filters
    if vessel_type_filter:
        specifications = specifications.filter(vessel_type=vessel_type_filter)
    if flag_filter:
        specifications = specifications.filter(flag=flag_filter)
    if coating_filter:
        specifications = specifications.filter(cargo_tank_coating=coating_filter)
    if fuel_type_filter:
        specifications = specifications.filter(fuel_type=fuel_type_filter)
    if search_query:
        specifications = specifications.filter(
            Q(vessel_name__icontains=search_query) |
            Q(imo_number__icontains=search_query) |
            Q(owner_name__icontains=search_query) |
            Q(classification_society__icontains=search_query)
        )

    # Filter by TC status
    if tc_status_filter == 'active':
        # Get IMOs with active TC contracts
        active_imos = TCFleet.objects.filter(delivery_status='ON_TC').values_list('imo_number', flat=True)
        specifications = specifications.filter(imo_number__in=active_imos)
    elif tc_status_filter == 'inactive':
        # Get IMOs with active TC contracts
        active_imos = TCFleet.objects.filter(delivery_status='ON_TC').values_list('imo_number', flat=True)
        specifications = specifications.exclude(imo_number__in=active_imos)

    # Get unique values for filter dropdowns
    vessel_types = ShipSpecification.objects.values_list('vessel_type', flat=True).distinct().order_by('vessel_type')
    flags = ShipSpecification.objects.values_list('flag', flat=True).distinct().order_by('flag')
    coatings = ShipSpecification.objects.values_list('cargo_tank_coating', flat=True).distinct().order_by('cargo_tank_coating')
    fuel_types = ShipSpecification.objects.values_list('fuel_type', flat=True).distinct().order_by('fuel_type')

    # Calculate statistics
    total_ships = specifications.count()
    crude_tankers = specifications.filter(vessel_type='CRUDE').count()
    product_tankers = specifications.filter(vessel_type='PRODUCT').count()
    chemical_tankers = specifications.filter(vessel_type='CHEMICAL').count()

    context = {
        'specifications': specifications,
        'vessel_types': vessel_types,
        'flags': flags,
        'coatings': coatings,
        'fuel_types': fuel_types,
        'vessel_type_filter': vessel_type_filter,
        'flag_filter': flag_filter,
        'coating_filter': coating_filter,
        'fuel_type_filter': fuel_type_filter,
        'tc_status_filter': tc_status_filter,
        'search_query': search_query,
        'total_ships': total_ships,
        'crude_tankers': crude_tankers,
        'product_tankers': product_tankers,
        'chemical_tankers': chemical_tankers,
    }

    return render(request, 'ships/ship_specifications_list.html', context)


@login_required
def ship_specification_detail(request, imo_number):
    """
    Display detailed view of a single ship specification.
    Shows all Q88 data organized in sections.
    """
    from ships.models import ShipSpecification

    specification = get_object_or_404(ShipSpecification, imo_number=imo_number)

    # Get all TC contracts for this vessel
    tc_contracts = specification.get_tc_fleet_contracts()
    active_contract = specification.get_active_tc_contract()

    context = {
        'spec': specification,
        'tc_contracts': tc_contracts,
        'active_contract': active_contract,
    }

    return render(request, 'ships/ship_specification_detail.html', context)


@login_required
def ship_specifications_export(request):
    """
    Export Ship Specifications to Excel.
    Applies the same filters as the list view.
    """
    from ships.models import ShipSpecification

    # Check if user has export permission
    if not request.user.can_export():
        return redirect('ship_specifications_list')

    # Get filter parameters (same as list view)
    vessel_type_filter = request.GET.get('vessel_type', '')
    flag_filter = request.GET.get('flag', '')
    coating_filter = request.GET.get('coating', '')
    fuel_type_filter = request.GET.get('fuel_type', '')
    tc_status_filter = request.GET.get('tc_status', '')
    search_query = request.GET.get('search', '')

    # Base queryset
    specifications = ShipSpecification.objects.all().select_related('created_by')

    # Apply filters
    if vessel_type_filter:
        specifications = specifications.filter(vessel_type=vessel_type_filter)
    if flag_filter:
        specifications = specifications.filter(flag=flag_filter)
    if coating_filter:
        specifications = specifications.filter(cargo_tank_coating=coating_filter)
    if fuel_type_filter:
        specifications = specifications.filter(fuel_type=fuel_type_filter)
    if search_query:
        specifications = specifications.filter(
            Q(vessel_name__icontains=search_query) |
            Q(imo_number__icontains=search_query) |
            Q(owner_name__icontains=search_query) |
            Q(classification_society__icontains=search_query)
        )

    # Filter by TC status
    if tc_status_filter == 'active':
        active_imos = TCFleet.objects.filter(delivery_status='ON_TC').values_list('imo_number', flat=True)
        specifications = specifications.filter(imo_number__in=active_imos)
    elif tc_status_filter == 'inactive':
        active_imos = TCFleet.objects.filter(delivery_status='ON_TC').values_list('imo_number', flat=True)
        specifications = specifications.exclude(imo_number__in=active_imos)

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ship Specifications Q88"

    # Define headers (all Q88 fields)
    headers = [
        # Vessel Identification
        'Vessel Name', 'IMO Number', 'Call Sign', 'Flag', 'Port of Registry',
        'Official Number', 'Vessel Type', 'Built Year', 'Built Country', 'Shipyard',
        'Classification Society', 'Class Notation',
        # Dimensions & Tonnages
        'LOA (m)', 'LBP (m)', 'Breadth (m)', 'Depth (m)', 'Summer Draft (m)',
        'Summer DWT (MT)', 'Lightweight (MT)', 'Gross Tonnage', 'Net Tonnage',
        'Suez Canal Tonnage', 'Panama Canal Tonnage', 'Displacement (MT)',
        # Cargo Capacity
        'Total Cargo Capacity (m³)', 'Number of Cargo Tanks', 'Avg Capacity per Tank (m³)',
        'Segregated Ballast Capacity (m³)', 'Slop Tank Capacity (m³)',
        'Cargo Tank Coating', 'Cargo Heating', 'Max Heating Temp (°C)',
        # Machinery & Performance
        'Main Engine Type', 'Main Engine Power (kW)', 'Engine Builder',
        'Service Speed Laden (kts)', 'Service Speed Ballast (kts)',
        'Fuel Consumption Laden (T/day)', 'Fuel Consumption Ballast (T/day)',
        'Fuel Type', 'Bow Thruster', 'Stern Thruster',
        # Cargo Handling
        'Number of Cargo Pumps', 'Cargo Pump Capacity (m³/h)', 'Inert Gas System',
        'Crude Oil Washing', 'Vapor Recovery System', 'Cargo Manifold Size (in)',
        'Manifold Pressure Rating',
        # Environmental & Safety
        'Double Hull', 'Ice Class', 'IOPP Cert Expiry', 'SMC Expiry',
        'Safety Equipment Cert Expiry', 'Int. Oil Pollution Cert Expiry',
        'Ship Sanitation Cert Expiry',
        # Operational Requirements
        'Min Freeboard Laden (m)', 'Air Draft Ballast (m)', 'Air Draft Laden (m)',
        'Max Draft Restriction (m)', 'Port Restrictions', 'Special Requirements',
        # Commercial
        'Owner Name', 'Operator Name', 'Commercial Manager', 'Technical Manager',
        'P&I Club',
        # TC Status
        'TC Status', 'Active Contract Until'
    ]

    # Style for header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data rows
    for row_num, spec in enumerate(specifications, 2):
        col = 1

        # Vessel Identification
        ws.cell(row=row_num, column=col).value = spec.vessel_name; col += 1
        ws.cell(row=row_num, column=col).value = spec.imo_number; col += 1
        ws.cell(row=row_num, column=col).value = spec.call_sign; col += 1
        ws.cell(row=row_num, column=col).value = spec.flag; col += 1
        ws.cell(row=row_num, column=col).value = spec.port_of_registry; col += 1
        ws.cell(row=row_num, column=col).value = spec.official_number; col += 1
        ws.cell(row=row_num, column=col).value = spec.get_vessel_type_display(); col += 1
        ws.cell(row=row_num, column=col).value = spec.built_year; col += 1
        ws.cell(row=row_num, column=col).value = spec.built_country; col += 1
        ws.cell(row=row_num, column=col).value = spec.shipyard; col += 1
        ws.cell(row=row_num, column=col).value = spec.classification_society; col += 1
        ws.cell(row=row_num, column=col).value = spec.class_notation; col += 1

        # Dimensions & Tonnages
        ws.cell(row=row_num, column=col).value = float(spec.length_overall); col += 1
        ws.cell(row=row_num, column=col).value = float(spec.length_between_perpendiculars) if spec.length_between_perpendiculars else None; col += 1
        ws.cell(row=row_num, column=col).value = float(spec.breadth_moulded); col += 1
        ws.cell(row=row_num, column=col).value = float(spec.depth_moulded); col += 1
        ws.cell(row=row_num, column=col).value = float(spec.summer_draft); col += 1
        ws.cell(row=row_num, column=col).value = float(spec.summer_deadweight); col += 1
        ws.cell(row=row_num, column=col).value = float(spec.lightweight); col += 1
        ws.cell(row=row_num, column=col).value = float(spec.gross_tonnage); col += 1
        ws.cell(row=row_num, column=col).value = float(spec.net_tonnage); col += 1
        ws.cell(row=row_num, column=col).value = float(spec.suez_canal_tonnage) if spec.suez_canal_tonnage else None; col += 1
        ws.cell(row=row_num, column=col).value = float(spec.panama_canal_tonnage) if spec.panama_canal_tonnage else None; col += 1
        ws.cell(row=row_num, column=col).value = float(spec.displacement); col += 1

        # Cargo Capacity
        ws.cell(row=row_num, column=col).value = float(spec.total_cargo_capacity); col += 1
        ws.cell(row=row_num, column=col).value = spec.number_of_cargo_tanks; col += 1
        ws.cell(row=row_num, column=col).value = float(spec.cargo_capacity_per_tank); col += 1
        ws.cell(row=row_num, column=col).value = float(spec.segregated_ballast_tanks_capacity) if spec.segregated_ballast_tanks_capacity else None; col += 1
        ws.cell(row=row_num, column=col).value = float(spec.slop_tank_capacity) if spec.slop_tank_capacity else None; col += 1
        ws.cell(row=row_num, column=col).value = spec.get_cargo_tank_coating_display(); col += 1
        ws.cell(row=row_num, column=col).value = 'Yes' if spec.cargo_heating_capability else 'No'; col += 1
        ws.cell(row=row_num, column=col).value = float(spec.maximum_heating_temperature) if spec.maximum_heating_temperature else None; col += 1

        # Machinery & Performance
        ws.cell(row=row_num, column=col).value = spec.main_engine_type; col += 1
        ws.cell(row=row_num, column=col).value = float(spec.main_engine_power); col += 1
        ws.cell(row=row_num, column=col).value = spec.main_engine_builder; col += 1
        ws.cell(row=row_num, column=col).value = float(spec.service_speed_laden); col += 1
        ws.cell(row=row_num, column=col).value = float(spec.service_speed_ballast); col += 1
        ws.cell(row=row_num, column=col).value = float(spec.fuel_consumption_laden); col += 1
        ws.cell(row=row_num, column=col).value = float(spec.fuel_consumption_ballast); col += 1
        ws.cell(row=row_num, column=col).value = spec.get_fuel_type_display(); col += 1
        ws.cell(row=row_num, column=col).value = 'Yes' if spec.bow_thruster else 'No'; col += 1
        ws.cell(row=row_num, column=col).value = 'Yes' if spec.stern_thruster else 'No'; col += 1

        # Cargo Handling
        ws.cell(row=row_num, column=col).value = spec.number_of_cargo_pumps; col += 1
        ws.cell(row=row_num, column=col).value = float(spec.cargo_pump_capacity); col += 1
        ws.cell(row=row_num, column=col).value = 'Yes' if spec.inert_gas_system else 'No'; col += 1
        ws.cell(row=row_num, column=col).value = 'Yes' if spec.crude_oil_washing else 'No'; col += 1
        ws.cell(row=row_num, column=col).value = 'Yes' if spec.vapor_recovery_system else 'No'; col += 1
        ws.cell(row=row_num, column=col).value = float(spec.cargo_manifold_size); col += 1
        ws.cell(row=row_num, column=col).value = float(spec.cargo_manifold_pressure_rating) if spec.cargo_manifold_pressure_rating else None; col += 1

        # Environmental & Safety
        ws.cell(row=row_num, column=col).value = 'Yes' if spec.double_hull else 'No'; col += 1
        ws.cell(row=row_num, column=col).value = spec.ice_class; col += 1
        ws.cell(row=row_num, column=col).value = spec.oil_pollution_prevention_certificate_expiry.strftime('%d/%m/%Y') if spec.oil_pollution_prevention_certificate_expiry else ''; col += 1
        ws.cell(row=row_num, column=col).value = spec.safety_management_certificate_expiry.strftime('%d/%m/%Y') if spec.safety_management_certificate_expiry else ''; col += 1
        ws.cell(row=row_num, column=col).value = spec.safety_equipment_certificate_expiry.strftime('%d/%m/%Y') if spec.safety_equipment_certificate_expiry else ''; col += 1
        ws.cell(row=row_num, column=col).value = spec.international_oil_pollution_certificate_expiry.strftime('%d/%m/%Y') if spec.international_oil_pollution_certificate_expiry else ''; col += 1
        ws.cell(row=row_num, column=col).value = spec.ship_sanitation_certificate_expiry.strftime('%d/%m/%Y') if spec.ship_sanitation_certificate_expiry else ''; col += 1

        # Operational Requirements
        ws.cell(row=row_num, column=col).value = float(spec.minimum_freeboard_laden) if spec.minimum_freeboard_laden else None; col += 1
        ws.cell(row=row_num, column=col).value = float(spec.air_draft_ballast) if spec.air_draft_ballast else None; col += 1
        ws.cell(row=row_num, column=col).value = float(spec.air_draft_laden) if spec.air_draft_laden else None; col += 1
        ws.cell(row=row_num, column=col).value = float(spec.maximum_allowed_draft_restriction) if spec.maximum_allowed_draft_restriction else None; col += 1
        ws.cell(row=row_num, column=col).value = spec.port_restrictions; col += 1
        ws.cell(row=row_num, column=col).value = spec.special_requirements; col += 1

        # Commercial
        ws.cell(row=row_num, column=col).value = spec.owner_name; col += 1
        ws.cell(row=row_num, column=col).value = spec.operator_name; col += 1
        ws.cell(row=row_num, column=col).value = spec.commercial_manager; col += 1
        ws.cell(row=row_num, column=col).value = spec.technical_manager; col += 1
        ws.cell(row=row_num, column=col).value = spec.p_and_i_club; col += 1

        # TC Status
        active_contract = spec.get_active_tc_contract()
        ws.cell(row=row_num, column=col).value = 'On TC' if active_contract else 'No Active Contract'; col += 1
        ws.cell(row=row_num, column=col).value = active_contract.redelivery_date.strftime('%d/%m/%Y') if active_contract else ''; col += 1

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=ship_specifications_q88_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    wb.save(response)
    return response
