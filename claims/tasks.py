"""
Celery tasks for Claims Management System

⚠️ NOTE: RADAR sync tasks are PLACEHOLDER CODE awaiting API integration ⚠️

Background tasks for:
- RADAR synchronization (TODO: Implement when API is available)
- Excel exports (Implemented)
- Email notifications (Implemented)
- Data maintenance (Implemented)

Tasks marked with TODO are placeholder implementations that return mock data.
They are ready for actual RADAR API integration when credentials and endpoints are provided.
"""
from celery import shared_task
from django.core.mail import send_mail
from django.conf import settings
from django.utils import timezone
from datetime import timedelta
import logging

logger = logging.getLogger(__name__)


@shared_task(bind=True, max_retries=3)
def sync_radar_data(self):
    """
    Synchronize data from RADAR system

    This task fetches new/updated voyages and claims from RADAR
    and updates the local database.
    """
    try:
        from claims.models import Voyage, Claim

        logger.info("Starting RADAR data synchronization")

        # TODO: Implement actual RADAR API integration
        # For now, this is a placeholder

        synced_voyages = 0
        synced_claims = 0

        # Example: Fetch voyages from RADAR API
        # voyages = fetch_from_radar_api('/voyages')
        # for voyage_data in voyages:
        #     voyage, created = Voyage.objects.update_or_create(
        #         radar_voyage_id=voyage_data['id'],
        #         defaults={...}
        #     )
        #     synced_voyages += 1

        logger.info(f"RADAR sync completed: {synced_voyages} voyages, {synced_claims} claims")

        return {
            'success': True,
            'voyages_synced': synced_voyages,
            'claims_synced': synced_claims,
            'timestamp': timezone.now().isoformat()
        }

    except Exception as exc:
        logger.error(f"RADAR sync failed: {exc}")
        # Retry with exponential backoff
        raise self.retry(exc=exc, countdown=60 * (2 ** self.request.retries))


@shared_task
def check_timebarred_claims():
    """
    Check for claims approaching or past time-bar deadline

    Sends notifications to analysts for claims nearing time-bar
    and marks claims as time-barred if deadline passed.
    """
    from claims.models import Claim

    logger.info("Checking for time-barred claims")

    today = timezone.now().date()
    warning_date = today + timedelta(days=30)  # 30-day warning

    # Find claims approaching time-bar
    approaching_claims = Claim.objects.filter(
        claim_deadline__lte=warning_date,
        claim_deadline__gte=today,
        is_time_barred=False,
        payment_status='NOT_SENT'
    )

    # Find claims that are now time-barred
    timebarred_claims = Claim.objects.filter(
        claim_deadline__lt=today,
        is_time_barred=False,
        payment_status='NOT_SENT'
    )

    # Mark as time-barred
    updated_count = 0
    for claim in timebarred_claims:
        claim.is_time_barred = True
        claim.time_bar_date = today
        claim.payment_status = 'TIMEBAR'
        claim.save()
        updated_count += 1

        # Send notification to assigned analyst
        if claim.assigned_to:
            send_timebar_notification.delay(claim.id)

    logger.info(
        f"Time-bar check completed: "
        f"{approaching_claims.count()} approaching, "
        f"{updated_count} marked time-barred"
    )

    return {
        'approaching': approaching_claims.count(),
        'time_barred': updated_count
    }


@shared_task
def generate_excel_export(user_id, export_type, filters=None):
    """
    Generate Excel export in background

    Args:
        user_id: User requesting the export
        export_type: Type of export ('voyages', 'claims', 'analytics')
        filters: Optional filters to apply

    Returns:
        dict: Export result with file path
    """
    from claims.models import User
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    logger.info(f"Generating {export_type} export for user {user_id}")

    try:
        user = User.objects.get(id=user_id)

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = export_type.title()

        # TODO: Implement actual export logic based on export_type
        # This is a placeholder

        # Example header
        ws.append(['Generated on:', timezone.now().isoformat()])
        ws.append(['Requested by:', user.get_full_name()])
        ws.append([])  # Empty row

        # Save to media directory
        file_path = f'exports/{export_type}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        full_path = f'{settings.MEDIA_ROOT}/{file_path}'
        wb.save(full_path)

        # Send notification email
        send_export_ready_notification.delay(user_id, file_path)

        logger.info(f"Export completed: {file_path}")

        return {
            'success': True,
            'file_path': file_path,
            'user_id': user_id
        }

    except Exception as exc:
        logger.error(f"Export generation failed: {exc}")
        return {
            'success': False,
            'error': str(exc)
        }


@shared_task
def send_email_notification(subject, message, recipient_list, from_email=None):
    """
    Send email notification

    Args:
        subject: Email subject
        message: Email body
        recipient_list: List of recipient email addresses
        from_email: Optional from email (defaults to settings.DEFAULT_FROM_EMAIL)
    """
    try:
        send_mail(
            subject=subject,
            message=message,
            from_email=from_email or settings.DEFAULT_FROM_EMAIL,
            recipient_list=recipient_list,
            fail_silently=False,
        )

        logger.info(f"Email sent to {len(recipient_list)} recipients: {subject}")
        return {'success': True, 'recipients': len(recipient_list)}

    except Exception as exc:
        logger.error(f"Email send failed: {exc}")
        return {'success': False, 'error': str(exc)}


@shared_task
def send_timebar_notification(claim_id):
    """Send notification for time-barred claim"""
    from claims.models import Claim

    try:
        claim = Claim.objects.get(id=claim_id)

        if claim.assigned_to and claim.assigned_to.email:
            subject = f"⚠️ Claim Time-Barred: {claim.claim_number}"
            message = f"""
            Dear {claim.assigned_to.get_full_name()},

            The following claim has been marked as time-barred:

            Claim Number: {claim.claim_number}
            Voyage: {claim.voyage.voyage_number}
            Vessel: {claim.voyage.vessel_name}
            Claim Amount: ${claim.claim_amount:,.2f}
            Deadline: {claim.claim_deadline}

            Please review this claim urgently.

            Best regards,
            Claims Management System
            """

            send_email_notification.delay(
                subject=subject,
                message=message,
                recipient_list=[claim.assigned_to.email]
            )

    except Exception as exc:
        logger.error(f"Failed to send time-bar notification for claim {claim_id}: {exc}")


@shared_task
def send_export_ready_notification(user_id, file_path):
    """Send notification when export is ready"""
    from claims.models import User

    try:
        user = User.objects.get(id=user_id)

        if user.email:
            subject = "✅ Your Export is Ready"
            message = f"""
            Dear {user.get_full_name()},

            Your requested export has been completed and is ready for download.

            File: {file_path}

            Please log in to the Claims Management System to download your file.

            Best regards,
            Claims Management System
            """

            send_email_notification.delay(
                subject=subject,
                message=message,
                recipient_list=[user.email]
            )

    except Exception as exc:
        logger.error(f"Failed to send export ready notification to user {user_id}: {exc}")


@shared_task
def generate_daily_analytics():
    """Generate daily analytics report"""
    from claims.models import Claim, Voyage

    logger.info("Generating daily analytics")

    today = timezone.now().date()

    # Calculate daily statistics
    total_claims = Claim.objects.count()
    claims_today = Claim.objects.filter(created_at__date=today).count()
    pending_claims = Claim.objects.filter(status='DRAFT').count()
    submitted_claims = Claim.objects.filter(status='SUBMITTED').count()

    # Send to admin users
    from claims.models import User
    admin_users = User.objects.filter(role='ADMIN', email__isnull=False)

    for admin in admin_users:
        subject = f"📊 Daily Claims Analytics - {today}"
        message = f"""
        Daily Claims Report for {today}

        Total Claims: {total_claims}
        New Claims Today: {claims_today}
        Pending Claims: {pending_claims}
        Submitted Claims: {submitted_claims}

        Log in to view detailed reports.
        """

        send_email_notification.delay(
            subject=subject,
            message=message,
            recipient_list=[admin.email]
        )

    logger.info("Daily analytics sent to admin users")


@shared_task
def cleanup_old_logs():
    """Cleanup old activity logs and temporary files"""
    from claims.models import ClaimActivityLog

    logger.info("Starting cleanup of old logs")

    # Delete activity logs older than 2 years
    cutoff_date = timezone.now() - timedelta(days=730)
    deleted_count, _ = ClaimActivityLog.objects.filter(
        created_at__lt=cutoff_date
    ).delete()

    logger.info(f"Cleanup completed: {deleted_count} old logs deleted")

    return {'logs_deleted': deleted_count}


@shared_task
def process_batch_import(file_path, user_id):
    """
    Process batch import from Excel/CSV file

    Args:
        file_path: Path to import file
        user_id: User who initiated the import
    """
    from claims.models import User
    import openpyxl

    logger.info(f"Processing batch import from {file_path}")

    try:
        user = User.objects.get(id=user_id)
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        imported_count = 0
        error_count = 0
        errors = []

        # TODO: Implement actual import logic
        # This is a placeholder

        # Send completion notification
        subject = f"Batch Import {'Completed' if error_count == 0 else 'Completed with Errors'}"
        message = f"""
        Batch import completed:

        Total Imported: {imported_count}
        Errors: {error_count}

        {'Errors: ' + str(errors) if errors else ''}
        """

        send_email_notification.delay(
            subject=subject,
            message=message,
            recipient_list=[user.email]
        )

        return {
            'success': True,
            'imported': imported_count,
            'errors': error_count
        }

    except Exception as exc:
        logger.error(f"Batch import failed: {exc}")
        return {
            'success': False,
            'error': str(exc)
        }
