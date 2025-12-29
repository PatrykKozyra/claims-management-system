from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse
from django.contrib import messages
from datetime import datetime
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .models import PortActivity, ActivityType
from ships.models import Ship
from claims.models import Voyage


# ============================================
# ACTIVITY LIST VIEWS
# ============================================

@login_required
def activity_list_all(request):
    """
    Display all port activities in timeline format.
    Shows activities grouped by ship + voyage with filters.
    """
    # Get filter parameters
    ship_filter = request.GET.get('ship', '')
    voyage_filter = request.GET.get('voyage', '')
    activity_type_filter = request.GET.get('activity_type', '')
    category_filter = request.GET.get('category', '')
    date_status_filter = request.GET.get('date_status', '')
    start_date_from = request.GET.get('start_date_from', '')
    start_date_to = request.GET.get('start_date_to', '')
    search_query = request.GET.get('search', '')

    # Base queryset with optimized select_related
    activities = PortActivity.objects.select_related(
        'ship',
        'voyage',
        'activity_type',
        'created_by'
    ).all()

    # Apply filters
    if ship_filter:
        activities = activities.filter(ship__imo_number=ship_filter)

    if voyage_filter:
        activities = activities.filter(voyage__id=voyage_filter)

    if activity_type_filter:
        activities = activities.filter(activity_type__code=activity_type_filter)

    if category_filter:
        activities = activities.filter(activity_type__category=category_filter)

    if date_status_filter:
        if date_status_filter == 'ACTUAL':
            activities = activities.filter(
                start_date_status='ACTUAL',
                end_date_status='ACTUAL'
            )
        elif date_status_filter == 'ESTIMATED':
            activities = activities.filter(
                Q(start_date_status='ESTIMATED') | Q(end_date_status='ESTIMATED')
            )

    if start_date_from:
        activities = activities.filter(start_datetime__gte=start_date_from)

    if start_date_to:
        activities = activities.filter(start_datetime__lte=start_date_to)

    if search_query:
        activities = activities.filter(
            Q(ship__vessel_name__icontains=search_query) |
            Q(ship__imo_number__icontains=search_query) |
            Q(port_name__icontains=search_query) |
            Q(voyage__voyage_number__icontains=search_query) |
            Q(notes__icontains=search_query)
        )

    # Order by ship, voyage, then start datetime
    activities = activities.order_by('ship__vessel_name', 'voyage__id', 'start_datetime')

    # Calculate statistics
    total_activities = activities.count()
    actual_activities = activities.filter(
        start_date_status='ACTUAL',
        end_date_status='ACTUAL'
    ).count()
    estimated_activities = total_activities - actual_activities

    # Get unique values for filter dropdowns
    ships = Ship.objects.filter(
        port_activities__isnull=False
    ).distinct().order_by('vessel_name')

    voyages = Voyage.objects.filter(
        port_activities__isnull=False
    ).distinct().select_related('ship_owner').order_by('-laycan_start')

    activity_types = ActivityType.objects.filter(is_active=True).order_by('category', 'name')

    categories = ActivityType.CATEGORY_CHOICES

    context = {
        'activities': activities,
        'ships': ships,
        'voyages': voyages,
        'activity_types': activity_types,
        'categories': categories,
        'ship_filter': ship_filter,
        'voyage_filter': voyage_filter,
        'activity_type_filter': activity_type_filter,
        'category_filter': category_filter,
        'date_status_filter': date_status_filter,
        'start_date_from': start_date_from,
        'start_date_to': start_date_to,
        'search_query': search_query,
        'total_activities': total_activities,
        'actual_activities': actual_activities,
        'estimated_activities': estimated_activities,
    }

    return render(request, 'port_activities/activity_list_all.html', context)


@login_required
def activity_list_by_ship(request, imo_number):
    """Display activities for a specific ship"""
    ship = get_object_or_404(Ship, imo_number=imo_number)

    activities = PortActivity.objects.filter(
        ship=ship
    ).select_related('activity_type', 'voyage', 'created_by').order_by('-start_datetime')

    context = {
        'ship': ship,
        'activities': activities,
    }

    return render(request, 'port_activities/activity_list_by_ship.html', context)


@login_required
def activity_list_by_voyage(request, voyage_id):
    """Display activities for a specific voyage"""
    voyage = get_object_or_404(Voyage, id=voyage_id)

    activities = PortActivity.objects.filter(
        voyage=voyage
    ).select_related('activity_type', 'ship', 'created_by').order_by('start_datetime')

    context = {
        'voyage': voyage,
        'activities': activities,
    }

    return render(request, 'port_activities/activity_list_by_voyage.html', context)


# ============================================
# ACTIVITY CRUD VIEWS
# ============================================

@login_required
def activity_create(request):
    """Create a new port activity"""
    if request.method == 'POST':
        # Get form data
        ship_imo = request.POST.get('ship')
        voyage_id = request.POST.get('voyage')
        activity_type_code = request.POST.get('activity_type')
        port_name = request.POST.get('port_name')
        load_port = request.POST.get('load_port', '')
        discharge_port = request.POST.get('discharge_port', '')
        start_datetime = request.POST.get('start_datetime')
        start_date_status = request.POST.get('start_date_status')
        end_datetime = request.POST.get('end_datetime')
        end_date_status = request.POST.get('end_date_status')
        cargo_quantity = request.POST.get('cargo_quantity')
        notes = request.POST.get('notes', '')

        # Get related objects
        ship = get_object_or_404(Ship, imo_number=ship_imo)
        activity_type = get_object_or_404(ActivityType, code=activity_type_code)
        voyage = get_object_or_404(Voyage, id=voyage_id) if voyage_id else None

        # Create activity
        activity = PortActivity.objects.create(
            ship=ship,
            voyage=voyage,
            activity_type=activity_type,
            port_name=port_name,
            load_port=load_port,
            discharge_port=discharge_port,
            start_datetime=start_datetime,
            start_date_status=start_date_status,
            end_datetime=end_datetime,
            end_date_status=end_date_status,
            cargo_quantity=Decimal(cargo_quantity) if cargo_quantity else None,
            notes=notes,
            created_by=request.user
        )

        messages.success(request, f'Activity "{activity.activity_type.name}" created successfully.')
        return redirect('port_activities:activity_list_all')

    # GET request - show form
    ships = Ship.objects.filter(is_active=True).order_by('vessel_name')
    voyages = Voyage.objects.all().select_related('ship_owner').order_by('-laycan_start')[:100]
    activity_types = ActivityType.objects.filter(is_active=True).order_by('category', 'name')

    context = {
        'ships': ships,
        'voyages': voyages,
        'activity_types': activity_types,
        'date_status_choices': PortActivity.DATE_STATUS_CHOICES,
    }

    return render(request, 'port_activities/activity_form.html', context)


@login_required
def activity_detail(request, pk):
    """View detailed information about a port activity"""
    activity = get_object_or_404(
        PortActivity.objects.select_related(
            'ship', 'voyage', 'activity_type', 'created_by'
        ),
        pk=pk
    )

    context = {
        'activity': activity,
    }

    return render(request, 'port_activities/activity_detail.html', context)


@login_required
def activity_update(request, pk):
    """Update an existing port activity"""
    activity = get_object_or_404(PortActivity, pk=pk)

    if not request.user.can_write():
        messages.error(request, 'You do not have permission to edit activities.')
        return redirect('port_activities:activity_detail', pk=pk)

    if request.method == 'POST':
        # Update activity fields
        activity.port_name = request.POST.get('port_name')
        activity.load_port = request.POST.get('load_port', '')
        activity.discharge_port = request.POST.get('discharge_port', '')
        activity.start_datetime = request.POST.get('start_datetime')
        activity.start_date_status = request.POST.get('start_date_status')
        activity.end_datetime = request.POST.get('end_datetime')
        activity.end_date_status = request.POST.get('end_date_status')

        cargo_quantity = request.POST.get('cargo_quantity')
        activity.cargo_quantity = Decimal(cargo_quantity) if cargo_quantity else None
        activity.notes = request.POST.get('notes', '')

        activity.save()

        messages.success(request, 'Activity updated successfully.')
        return redirect('port_activities:activity_detail', pk=pk)

    # GET request - show form
    context = {
        'activity': activity,
        'date_status_choices': PortActivity.DATE_STATUS_CHOICES,
    }

    return render(request, 'port_activities/activity_edit.html', context)


@login_required
def activity_delete(request, pk):
    """Delete a port activity"""
    activity = get_object_or_404(PortActivity, pk=pk)

    if not request.user.can_write():
        messages.error(request, 'You do not have permission to delete activities.')
        return redirect('port_activities:activity_detail', pk=pk)

    if request.method == 'POST':
        activity.delete()
        messages.success(request, 'Activity deleted successfully.')
        return redirect('port_activities:activity_list_all')

    context = {
        'activity': activity,
    }

    return render(request, 'port_activities/activity_confirm_delete.html', context)


# ============================================
# EXPORT VIEW
# ============================================

@login_required
def activity_export(request):
    """Export port activities to Excel"""
    if not request.user.can_export():
        messages.error(request, 'You do not have permission to export data.')
        return redirect('port_activities:activity_list_all')

    # Apply same filters as list view
    activities = PortActivity.objects.select_related(
        'ship', 'voyage', 'activity_type', 'created_by'
    ).order_by('ship__vessel_name', 'voyage__id', 'start_datetime')

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Port Activities"

    # Define headers
    headers = [
        'Ship Name', 'IMO Number', 'Voyage Number', 'Activity Type', 'Category',
        'Port Name', 'Load Port', 'Discharge Port',
        'Start Date/Time', 'Start Status', 'End Date/Time', 'End Status',
        'Duration (Hours)', 'Duration (Days)',
        'Cargo Quantity (MT)', 'Notes', 'Created By', 'Created At'
    ]

    # Style for header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data rows
    for row_num, activity in enumerate(activities, 2):
        ws.cell(row=row_num, column=1).value = activity.ship.vessel_name
        ws.cell(row=row_num, column=2).value = activity.ship.imo_number
        ws.cell(row=row_num, column=3).value = activity.voyage.voyage_number if activity.voyage else ''
        ws.cell(row=row_num, column=4).value = activity.activity_type.name
        ws.cell(row=row_num, column=5).value = activity.activity_type.get_category_display()
        ws.cell(row=row_num, column=6).value = activity.port_name
        ws.cell(row=row_num, column=7).value = activity.load_port
        ws.cell(row=row_num, column=8).value = activity.discharge_port
        ws.cell(row=row_num, column=9).value = activity.start_datetime.strftime('%Y-%m-%d %H:%M')
        ws.cell(row=row_num, column=10).value = activity.get_start_date_status_display()
        ws.cell(row=row_num, column=11).value = activity.end_datetime.strftime('%Y-%m-%d %H:%M')
        ws.cell(row=row_num, column=12).value = activity.get_end_date_status_display()
        ws.cell(row=row_num, column=13).value = round(activity.duration_hours, 2)
        ws.cell(row=row_num, column=14).value = activity.duration_days
        ws.cell(row=row_num, column=15).value = float(activity.cargo_quantity) if activity.cargo_quantity else ''
        ws.cell(row=row_num, column=16).value = activity.notes
        ws.cell(row=row_num, column=17).value = activity.created_by.get_full_name() or activity.created_by.username
        ws.cell(row=row_num, column=18).value = activity.created_at.strftime('%Y-%m-%d %H:%M')

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=port_activities_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    wb.save(response)
    return response
