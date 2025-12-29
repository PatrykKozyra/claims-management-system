from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Count, Sum, Avg
from django.http import HttpResponse
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

from .models import TCFleet, Ship


@login_required
def tc_fleet_list(request):
    """
    Display list of TC Fleet contracts with filters.
    """
    # Get filter parameters
    status_filter = request.GET.get('status', '')
    ship_type_filter = request.GET.get('ship_type', '')
    trade_filter = request.GET.get('trade', '')
    owner_filter = request.GET.get('owner', '')
    trader_filter = request.GET.get('trader', '')
    search_query = request.GET.get('search', '')

    # Base queryset
    contracts = TCFleet.objects.all().select_related('created_by')

    # Apply filters
    if status_filter:
        contracts = contracts.filter(delivery_status=status_filter)
    if ship_type_filter:
        contracts = contracts.filter(ship_type=ship_type_filter)
    if trade_filter:
        contracts = contracts.filter(trade=trade_filter)
    if owner_filter:
        contracts = contracts.filter(owner_name=owner_filter)
    if trader_filter:
        contracts = contracts.filter(trader_name=trader_filter)
    if search_query:
        contracts = contracts.filter(
            Q(ship_name__icontains=search_query) |
            Q(imo_number__icontains=search_query) |
            Q(radar_deal_number__icontains=search_query) |
            Q(owner_name__icontains=search_query)
        )

    # Get unique values for filter dropdowns
    ship_types = TCFleet.objects.values_list('ship_type', flat=True).distinct().order_by('ship_type')
    trades = TCFleet.objects.values_list('trade', flat=True).distinct().order_by('trade')
    owners = TCFleet.objects.values_list('owner_name', flat=True).distinct().order_by('owner_name')
    traders = TCFleet.objects.filter(trader_name__isnull=False).exclude(trader_name='').values_list('trader_name', flat=True).distinct().order_by('trader_name')

    # Calculate statistics
    total_contracts = contracts.count()
    active_contracts = contracts.filter(delivery_status='ON_TC').count()
    incoming_contracts = contracts.filter(delivery_status='INCOMING_TC').count()
    redelivered_contracts = contracts.filter(delivery_status='REDELIVERED').count()

    context = {
        'contracts': contracts,
        'ship_types': ship_types,
        'trades': trades,
        'owners': owners,
        'traders': traders,
        'status_filter': status_filter,
        'ship_type_filter': ship_type_filter,
        'trade_filter': trade_filter,
        'owner_filter': owner_filter,
        'trader_filter': trader_filter,
        'search_query': search_query,
        'total_contracts': total_contracts,
        'active_contracts': active_contracts,
        'incoming_contracts': incoming_contracts,
        'redelivered_contracts': redelivered_contracts,
    }

    return render(request, 'ships/tc_fleet_list.html', context)


@login_required
def tc_fleet_detail(request, pk):
    """
    Display detailed view of a single TC Fleet contract.
    """
    contract = get_object_or_404(TCFleet, pk=pk)

    # Get all contracts for the same ship (IMO)
    contract_history = TCFleet.objects.filter(
        imo_number=contract.imo_number
    ).exclude(pk=pk).order_by('-delivery_date')

    # Get ship master data if exists
    ship_master = contract.get_ship_master_data()

    context = {
        'contract': contract,
        'contract_history': contract_history,
        'ship_master': ship_master,
    }

    return render(request, 'ships/tc_fleet_detail.html', context)


@login_required
def tc_fleet_by_ship(request, imo_number):
    """
    Display all TC Fleet contracts for a specific ship (by IMO number).
    Shows contract history timeline.
    """
    # Get all contracts for this IMO
    contracts = TCFleet.objects.filter(
        imo_number=imo_number
    ).order_by('-delivery_date')

    if not contracts.exists():
        return redirect('ships:tc_fleet_list')

    # Get ship master data if exists
    ship_master = None
    try:
        ship_master = Ship.objects.get(imo_number=imo_number)
    except Ship.DoesNotExist:
        pass

    # Get the most recent contract for ship name
    latest_contract = contracts.first()

    # Calculate statistics for this ship
    total_contracts = contracts.count()
    active_contracts = contracts.filter(delivery_status='ON_TC').count()
    redelivered_contracts = contracts.filter(delivery_status='REDELIVERED').count()

    context = {
        'contracts': contracts,
        'imo_number': imo_number,
        'ship_name': latest_contract.ship_name,
        'ship_master': ship_master,
        'total_contracts': total_contracts,
        'active_contracts': active_contracts,
        'redelivered_contracts': redelivered_contracts,
    }

    return render(request, 'ships/tc_fleet_by_ship.html', context)


@login_required
def tc_fleet_export(request):
    """
    Export TC Fleet contracts to Excel.
    Applies the same filters as the list view.
    """
    # Check if user has export permission
    if not request.user.can_export():
        return redirect('ships:tc_fleet_list')

    # Get filter parameters (same as list view)
    status_filter = request.GET.get('status', '')
    ship_type_filter = request.GET.get('ship_type', '')
    trade_filter = request.GET.get('trade', '')
    owner_filter = request.GET.get('owner', '')
    trader_filter = request.GET.get('trader', '')
    search_query = request.GET.get('search', '')

    # Base queryset
    contracts = TCFleet.objects.all().select_related('created_by')

    # Apply filters
    if status_filter:
        contracts = contracts.filter(delivery_status=status_filter)
    if ship_type_filter:
        contracts = contracts.filter(ship_type=ship_type_filter)
    if trade_filter:
        contracts = contracts.filter(trade=trade_filter)
    if owner_filter:
        contracts = contracts.filter(owner_name=owner_filter)
    if trader_filter:
        contracts = contracts.filter(trader_name=trader_filter)
    if search_query:
        contracts = contracts.filter(
            Q(ship_name__icontains=search_query) |
            Q(imo_number__icontains=search_query) |
            Q(radar_deal_number__icontains=search_query) |
            Q(owner_name__icontains=search_query)
        )

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TC Fleet Contracts"

    # Define headers
    headers = [
        'Ship Name', 'IMO Number', 'Ship Type', 'Delivery Status', 'Trade',
        'Owner Name', 'Owner Email', 'Technical Manager', 'Tech Manager Email',
        'Charter Length (Years)', 'TC Rate Monthly (USD)', 'RADAR Deal Number',
        'Delivery Date', 'Redelivery Date', 'TCP Date', 'Broker Name',
        'Broker Email', 'Broker Commission (%)', 'Delivery Location',
        'Redelivery Location', 'Bunkers Policy', 'Summer DWT', 'Built Year',
        'Flag', 'Next Dry-Dock Date', 'Trader Name', 'Days Remaining',
        'Contract Status', 'Total Contract Value (USD)'
    ]

    # Style for header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data rows
    for row_num, contract in enumerate(contracts, 2):
        ws.cell(row=row_num, column=1).value = contract.ship_name
        ws.cell(row=row_num, column=2).value = contract.imo_number
        ws.cell(row=row_num, column=3).value = contract.get_ship_type_display()
        ws.cell(row=row_num, column=4).value = contract.get_delivery_status_display()
        ws.cell(row=row_num, column=5).value = contract.get_trade_display()
        ws.cell(row=row_num, column=6).value = contract.owner_name
        ws.cell(row=row_num, column=7).value = contract.owner_email
        ws.cell(row=row_num, column=8).value = contract.technical_manager
        ws.cell(row=row_num, column=9).value = contract.technical_manager_email
        ws.cell(row=row_num, column=10).value = float(contract.charter_length_years)
        ws.cell(row=row_num, column=11).value = float(contract.tc_rate_monthly)
        ws.cell(row=row_num, column=12).value = contract.radar_deal_number
        ws.cell(row=row_num, column=13).value = contract.delivery_date.strftime('%d/%m/%Y')
        ws.cell(row=row_num, column=14).value = contract.redelivery_date.strftime('%d/%m/%Y')
        ws.cell(row=row_num, column=15).value = contract.tcp_date.strftime('%d/%m/%Y')
        ws.cell(row=row_num, column=16).value = contract.broker_name
        ws.cell(row=row_num, column=17).value = contract.broker_email
        ws.cell(row=row_num, column=18).value = float(contract.broker_commission) if contract.broker_commission else None
        ws.cell(row=row_num, column=19).value = contract.delivery_location
        ws.cell(row=row_num, column=20).value = contract.redelivery_location
        ws.cell(row=row_num, column=21).value = contract.get_bunkers_policy_display() if contract.bunkers_policy else ''
        ws.cell(row=row_num, column=22).value = float(contract.summer_dwt)
        ws.cell(row=row_num, column=23).value = contract.built_year
        ws.cell(row=row_num, column=24).value = contract.flag
        ws.cell(row=row_num, column=25).value = contract.next_drydock_date.strftime('%d/%m/%Y') if contract.next_drydock_date else ''
        ws.cell(row=row_num, column=26).value = contract.trader_name
        ws.cell(row=row_num, column=27).value = contract.days_remaining
        ws.cell(row=row_num, column=28).value = contract.contract_status
        ws.cell(row=row_num, column=29).value = float(contract.total_contract_value)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=tc_fleet_contracts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    wb.save(response)
    return response
# ============================================
# SHIP SPECIFICATIONS (Q88) VIEWS
# ============================================

@login_required
def ship_specifications_list(request):
    """
    Display list of Ship Specifications (Q88) with filters.
    """
    from ships.models import ShipSpecification

    # Get filter parameters
    vessel_type_filter = request.GET.get('vessel_type', '')
    flag_filter = request.GET.get('flag', '')
    coating_filter = request.GET.get('coating', '')
    fuel_type_filter = request.GET.get('fuel_type', '')
    tc_status_filter = request.GET.get('tc_status', '')
    search_query = request.GET.get('search', '')

    # Base queryset
    specifications = ShipSpecification.objects.all().select_related('created_by')

    # Apply filters
    if vessel_type_filter:
        specifications = specifications.filter(vessel_type=vessel_type_filter)
    if flag_filter:
        specifications = specifications.filter(flag=flag_filter)
    if coating_filter:
        specifications = specifications.filter(cargo_tank_coating=coating_filter)
    if fuel_type_filter:
        specifications = specifications.filter(fuel_type=fuel_type_filter)
    if search_query:
        specifications = specifications.filter(
            Q(vessel_name__icontains=search_query) |
            Q(imo_number__icontains=search_query) |
            Q(owner_name__icontains=search_query) |
            Q(classification_society__icontains=search_query)
        )

    # Filter by TC status
    if tc_status_filter == 'active':
        # Get IMOs with active TC contracts
        active_imos = TCFleet.objects.filter(delivery_status='ON_TC').values_list('imo_number', flat=True)
        specifications = specifications.filter(imo_number__in=active_imos)
    elif tc_status_filter == 'inactive':
        # Get IMOs with active TC contracts
        active_imos = TCFleet.objects.filter(delivery_status='ON_TC').values_list('imo_number', flat=True)
        specifications = specifications.exclude(imo_number__in=active_imos)

    # Get unique values for filter dropdowns
    vessel_types = ShipSpecification.objects.values_list('vessel_type', flat=True).distinct().order_by('vessel_type')
    flags = ShipSpecification.objects.values_list('flag', flat=True).distinct().order_by('flag')
    coatings = ShipSpecification.objects.values_list('cargo_tank_coating', flat=True).distinct().order_by('cargo_tank_coating')
    fuel_types = ShipSpecification.objects.values_list('fuel_type', flat=True).distinct().order_by('fuel_type')

    # Calculate statistics
    total_ships = specifications.count()
    crude_tankers = specifications.filter(vessel_type='CRUDE').count()
    product_tankers = specifications.filter(vessel_type='PRODUCT').count()
    chemical_tankers = specifications.filter(vessel_type='CHEMICAL').count()

    context = {
        'specifications': specifications,
        'vessel_types': vessel_types,
        'flags': flags,
        'coatings': coatings,
        'fuel_types': fuel_types,
        'vessel_type_filter': vessel_type_filter,
        'flag_filter': flag_filter,
        'coating_filter': coating_filter,
        'fuel_type_filter': fuel_type_filter,
        'tc_status_filter': tc_status_filter,
        'search_query': search_query,
        'total_ships': total_ships,
        'crude_tankers': crude_tankers,
        'product_tankers': product_tankers,
        'chemical_tankers': chemical_tankers,
    }

    return render(request, 'ships/ship_specifications_list.html', context)


@login_required
def ship_specification_detail(request, imo_number):
    """
    Display detailed view of a single ship specification.
    Shows all Q88 data organized in sections.
    """
    from ships.models import ShipSpecification

    specification = get_object_or_404(ShipSpecification, imo_number=imo_number)

    # Get all TC contracts for this vessel
    tc_contracts = specification.get_tc_fleet_contracts()
    active_contract = specification.get_active_tc_contract()

    context = {
        'spec': specification,
        'tc_contracts': tc_contracts,
        'active_contract': active_contract,
    }

    return render(request, 'ships/ship_specification_detail.html', context)


@login_required
def ship_specifications_export(request):
    """
    Export Ship Specifications to Excel.
    Applies the same filters as the list view.
    """
    from ships.models import ShipSpecification

    # Check if user has export permission
    if not request.user.can_export():
        return redirect('ship_specifications_list')

    # Get filter parameters (same as list view)
    vessel_type_filter = request.GET.get('vessel_type', '')
    flag_filter = request.GET.get('flag', '')
    coating_filter = request.GET.get('coating', '')
    fuel_type_filter = request.GET.get('fuel_type', '')
    tc_status_filter = request.GET.get('tc_status', '')
    search_query = request.GET.get('search', '')

    # Base queryset
    specifications = ShipSpecification.objects.all().select_related('created_by')

    # Apply filters
    if vessel_type_filter:
        specifications = specifications.filter(vessel_type=vessel_type_filter)
    if flag_filter:
        specifications = specifications.filter(flag=flag_filter)
    if coating_filter:
        specifications = specifications.filter(cargo_tank_coating=coating_filter)
    if fuel_type_filter:
        specifications = specifications.filter(fuel_type=fuel_type_filter)
    if search_query:
        specifications = specifications.filter(
            Q(vessel_name__icontains=search_query) |
            Q(imo_number__icontains=search_query) |
            Q(owner_name__icontains=search_query) |
            Q(classification_society__icontains=search_query)
        )

    # Filter by TC status
    if tc_status_filter == 'active':
        active_imos = TCFleet.objects.filter(delivery_status='ON_TC').values_list('imo_number', flat=True)
        specifications = specifications.filter(imo_number__in=active_imos)
    elif tc_status_filter == 'inactive':
        active_imos = TCFleet.objects.filter(delivery_status='ON_TC').values_list('imo_number', flat=True)
        specifications = specifications.exclude(imo_number__in=active_imos)

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ship Specifications Q88"

    # Define headers (all Q88 fields)
    headers = [
        # Vessel Identification
        'Vessel Name', 'IMO Number', 'Call Sign', 'Flag', 'Port of Registry',
        'Official Number', 'Vessel Type', 'Built Year', 'Built Country', 'Shipyard',
        'Classification Society', 'Class Notation',
        # Dimensions & Tonnages
        'LOA (m)', 'LBP (m)', 'Breadth (m)', 'Depth (m)', 'Summer Draft (m)',
        'Summer DWT (MT)', 'Lightweight (MT)', 'Gross Tonnage', 'Net Tonnage',
        'Suez Canal Tonnage', 'Panama Canal Tonnage', 'Displacement (MT)',
        # Cargo Capacity
        'Total Cargo Capacity (m³)', 'Number of Cargo Tanks', 'Avg Capacity per Tank (m³)',
        'Segregated Ballast Capacity (m³)', 'Slop Tank Capacity (m³)',
        'Cargo Tank Coating', 'Cargo Heating', 'Max Heating Temp (°C)',
        # Machinery & Performance
        'Main Engine Type', 'Main Engine Power (kW)', 'Engine Builder',
        'Service Speed Laden (kts)', 'Service Speed Ballast (kts)',
        'Fuel Consumption Laden (T/day)', 'Fuel Consumption Ballast (T/day)',
        'Fuel Type', 'Bow Thruster', 'Stern Thruster',
        # Cargo Handling
        'Number of Cargo Pumps', 'Cargo Pump Capacity (m³/h)', 'Inert Gas System',
        'Crude Oil Washing', 'Vapor Recovery System', 'Cargo Manifold Size (in)',
        'Manifold Pressure Rating',
        # Environmental & Safety
        'Double Hull', 'Ice Class', 'IOPP Cert Expiry', 'SMC Expiry',
        'Safety Equipment Cert Expiry', 'Int. Oil Pollution Cert Expiry',
        'Ship Sanitation Cert Expiry',
        # Operational Requirements
        'Min Freeboard Laden (m)', 'Air Draft Ballast (m)', 'Air Draft Laden (m)',
        'Max Draft Restriction (m)', 'Port Restrictions', 'Special Requirements',
        # Commercial
        'Owner Name', 'Operator Name', 'Commercial Manager', 'Technical Manager',
        'P&I Club',
        # TC Status
        'TC Status', 'Active Contract Until'
    ]

    # Style for header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data rows
    for row_num, spec in enumerate(specifications, 2):
        col = 1

        # Vessel Identification
        ws.cell(row=row_num, column=col).value = spec.vessel_name; col += 1
        ws.cell(row=row_num, column=col).value = spec.imo_number; col += 1
        ws.cell(row=row_num, column=col).value = spec.call_sign; col += 1
        ws.cell(row=row_num, column=col).value = spec.flag; col += 1
        ws.cell(row=row_num, column=col).value = spec.port_of_registry; col += 1
        ws.cell(row=row_num, column=col).value = spec.official_number; col += 1
        ws.cell(row=row_num, column=col).value = spec.get_vessel_type_display(); col += 1
        ws.cell(row=row_num, column=col).value = spec.built_year; col += 1
        ws.cell(row=row_num, column=col).value = spec.built_country; col += 1
        ws.cell(row=row_num, column=col).value = spec.shipyard; col += 1
        ws.cell(row=row_num, column=col).value = spec.classification_society; col += 1
        ws.cell(row=row_num, column=col).value = spec.class_notation; col += 1

        # Dimensions & Tonnages
        ws.cell(row=row_num, column=col).value = float(spec.length_overall); col += 1
        ws.cell(row=row_num, column=col).value = float(spec.length_between_perpendiculars) if spec.length_between_perpendiculars else None; col += 1
        ws.cell(row=row_num, column=col).value = float(spec.breadth_moulded); col += 1
        ws.cell(row=row_num, column=col).value = float(spec.depth_moulded); col += 1
        ws.cell(row=row_num, column=col).value = float(spec.summer_draft); col += 1
        ws.cell(row=row_num, column=col).value = float(spec.summer_deadweight); col += 1
        ws.cell(row=row_num, column=col).value = float(spec.lightweight); col += 1
        ws.cell(row=row_num, column=col).value = float(spec.gross_tonnage); col += 1
        ws.cell(row=row_num, column=col).value = float(spec.net_tonnage); col += 1
        ws.cell(row=row_num, column=col).value = float(spec.suez_canal_tonnage) if spec.suez_canal_tonnage else None; col += 1
        ws.cell(row=row_num, column=col).value = float(spec.panama_canal_tonnage) if spec.panama_canal_tonnage else None; col += 1
        ws.cell(row=row_num, column=col).value = float(spec.displacement); col += 1

        # Cargo Capacity
        ws.cell(row=row_num, column=col).value = float(spec.total_cargo_capacity); col += 1
        ws.cell(row=row_num, column=col).value = spec.number_of_cargo_tanks; col += 1
        ws.cell(row=row_num, column=col).value = float(spec.cargo_capacity_per_tank); col += 1
        ws.cell(row=row_num, column=col).value = float(spec.segregated_ballast_tanks_capacity) if spec.segregated_ballast_tanks_capacity else None; col += 1
        ws.cell(row=row_num, column=col).value = float(spec.slop_tank_capacity) if spec.slop_tank_capacity else None; col += 1
        ws.cell(row=row_num, column=col).value = spec.get_cargo_tank_coating_display(); col += 1
        ws.cell(row=row_num, column=col).value = 'Yes' if spec.cargo_heating_capability else 'No'; col += 1
        ws.cell(row=row_num, column=col).value = float(spec.maximum_heating_temperature) if spec.maximum_heating_temperature else None; col += 1

        # Machinery & Performance
        ws.cell(row=row_num, column=col).value = spec.main_engine_type; col += 1
        ws.cell(row=row_num, column=col).value = float(spec.main_engine_power); col += 1
        ws.cell(row=row_num, column=col).value = spec.main_engine_builder; col += 1
        ws.cell(row=row_num, column=col).value = float(spec.service_speed_laden); col += 1
        ws.cell(row=row_num, column=col).value = float(spec.service_speed_ballast); col += 1
        ws.cell(row=row_num, column=col).value = float(spec.fuel_consumption_laden); col += 1
        ws.cell(row=row_num, column=col).value = float(spec.fuel_consumption_ballast); col += 1
        ws.cell(row=row_num, column=col).value = spec.get_fuel_type_display(); col += 1
        ws.cell(row=row_num, column=col).value = 'Yes' if spec.bow_thruster else 'No'; col += 1
        ws.cell(row=row_num, column=col).value = 'Yes' if spec.stern_thruster else 'No'; col += 1

        # Cargo Handling
        ws.cell(row=row_num, column=col).value = spec.number_of_cargo_pumps; col += 1
        ws.cell(row=row_num, column=col).value = float(spec.cargo_pump_capacity); col += 1
        ws.cell(row=row_num, column=col).value = 'Yes' if spec.inert_gas_system else 'No'; col += 1
        ws.cell(row=row_num, column=col).value = 'Yes' if spec.crude_oil_washing else 'No'; col += 1
        ws.cell(row=row_num, column=col).value = 'Yes' if spec.vapor_recovery_system else 'No'; col += 1
        ws.cell(row=row_num, column=col).value = float(spec.cargo_manifold_size); col += 1
        ws.cell(row=row_num, column=col).value = float(spec.cargo_manifold_pressure_rating) if spec.cargo_manifold_pressure_rating else None; col += 1

        # Environmental & Safety
        ws.cell(row=row_num, column=col).value = 'Yes' if spec.double_hull else 'No'; col += 1
        ws.cell(row=row_num, column=col).value = spec.ice_class; col += 1
        ws.cell(row=row_num, column=col).value = spec.oil_pollution_prevention_certificate_expiry.strftime('%d/%m/%Y') if spec.oil_pollution_prevention_certificate_expiry else ''; col += 1
        ws.cell(row=row_num, column=col).value = spec.safety_management_certificate_expiry.strftime('%d/%m/%Y') if spec.safety_management_certificate_expiry else ''; col += 1
        ws.cell(row=row_num, column=col).value = spec.safety_equipment_certificate_expiry.strftime('%d/%m/%Y') if spec.safety_equipment_certificate_expiry else ''; col += 1
        ws.cell(row=row_num, column=col).value = spec.international_oil_pollution_certificate_expiry.strftime('%d/%m/%Y') if spec.international_oil_pollution_certificate_expiry else ''; col += 1
        ws.cell(row=row_num, column=col).value = spec.ship_sanitation_certificate_expiry.strftime('%d/%m/%Y') if spec.ship_sanitation_certificate_expiry else ''; col += 1

        # Operational Requirements
        ws.cell(row=row_num, column=col).value = float(spec.minimum_freeboard_laden) if spec.minimum_freeboard_laden else None; col += 1
        ws.cell(row=row_num, column=col).value = float(spec.air_draft_ballast) if spec.air_draft_ballast else None; col += 1
        ws.cell(row=row_num, column=col).value = float(spec.air_draft_laden) if spec.air_draft_laden else None; col += 1
        ws.cell(row=row_num, column=col).value = float(spec.maximum_allowed_draft_restriction) if spec.maximum_allowed_draft_restriction else None; col += 1
        ws.cell(row=row_num, column=col).value = spec.port_restrictions; col += 1
        ws.cell(row=row_num, column=col).value = spec.special_requirements; col += 1

        # Commercial
        ws.cell(row=row_num, column=col).value = spec.owner_name; col += 1
        ws.cell(row=row_num, column=col).value = spec.operator_name; col += 1
        ws.cell(row=row_num, column=col).value = spec.commercial_manager; col += 1
        ws.cell(row=row_num, column=col).value = spec.technical_manager; col += 1
        ws.cell(row=row_num, column=col).value = spec.p_and_i_club; col += 1

        # TC Status
        active_contract = spec.get_active_tc_contract()
        ws.cell(row=row_num, column=col).value = 'On TC' if active_contract else 'No Active Contract'; col += 1
        ws.cell(row=row_num, column=col).value = active_contract.redelivery_date.strftime('%d/%m/%Y') if active_contract else ''; col += 1

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=ship_specifications_q88_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    wb.save(response)
    return response
